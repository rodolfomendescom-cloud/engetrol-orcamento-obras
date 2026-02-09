"""
Orçamentação de Obras — Engetrol Engenharia
Aplicação Streamlit (Wizard) — Regime Lucro Real
v3.0 — Histograma, Simulador de Negociação, Importação de Materiais, Lucro Líquido Alvo
"""

import streamlit as st
import pandas as pd
import uuid
import base64
import io
import os
import json
import math
from datetime import datetime, date, timedelta

# ══════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA
# ══════════════════════════════════════════════
st.set_page_config(
    page_title="Engetrol — Orçamento de Obras",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════
# LOGO EM BASE64
# ══════════════════════════════════════════════
LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo_horizontal.jpg")
LOGO_VERT_PATH = os.path.join(os.path.dirname(__file__), "logo_vertical.jpg")


def get_logo_base64(path):
    if os.path.exists(path):
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return ""


LOGO_B64 = get_logo_base64(LOGO_PATH)
LOGO_VERT_B64 = get_logo_base64(LOGO_VERT_PATH)

# ══════════════════════════════════════════════
# CSS — IDENTIDADE VISUAL ENGETROL v3.0
# ══════════════════════════════════════════════
st.markdown("""
<style>
.block-container {padding-top: 1.5rem; padding-bottom: 2rem;}
h1 {color: #1a1a1a; font-size: 1.8rem !important; font-weight: 700;}
h2 {color: #ed1c24; font-size: 1.35rem !important; border-bottom: 2px solid #ed1c24; padding-bottom: .3rem; font-weight: 700;}
h3 {color: #333; font-size: 1.1rem !important; font-weight: 600;}
div[data-testid="stMetric"] {
    background: linear-gradient(135deg, #ffffff 0%, #f1f2f2 100%);
    border: 1px solid #dee2e6; border-left: 4px solid #ed1c24;
    border-radius: 8px; padding: 14px 18px;
    box-shadow: 0 2px 6px rgba(0,0,0,.06);
}
div[data-testid="stMetric"] label {font-size: .82rem !important; color: #555; font-weight: 600;}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {font-size: 1.4rem !important; color: #1a1a1a; font-weight: 700;}
section[data-testid="stSidebar"] {background: linear-gradient(180deg, #1a1a1a 0%, #2d2d2d 100%);}
section[data-testid="stSidebar"] .stRadio label p,
section[data-testid="stSidebar"] h1, section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3, section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span, section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stMarkdown {
    color: #ffffff !important;
    font-weight: bold !important;
}
section[data-testid="stSidebar"] .stRadio label p {
    font-size: 1.1rem !important;
}
section[data-testid="stSidebar"] img {
    max-width: 180px;
    margin-bottom: 20px;
    border-radius: 10px;
}
.stButton>button {border-radius: 6px; font-weight: 600; transition: all .2s;}
.stButton>button[kind="primary"], div.stButton>button:first-child {
    background: linear-gradient(135deg, #ed1c24 0%, #a01e21 100%) !important;
    color: white !important; border: none !important;
}
.stButton>button:hover {transform: translateY(-1px); box-shadow: 0 4px 12px rgba(237,28,36,.3);}
.stDataFrame {border-radius: 6px; overflow: hidden;}
.streamlit-expanderHeader {font-weight: 600; color: #1a1a1a;}
hr {border-color: #e0e0e0 !important;}
.stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {
    border-bottom-color: #ed1c24 !important; color: #ed1c24 !important;
}
div[data-testid="stAlert"] {border-radius: 6px;}
.alerta-consistencia {
    background: #fff3cd; border: 1px solid #ffc107; border-left: 4px solid #ff6600;
    border-radius: 6px; padding: 12px 16px; margin: 8px 0;
    font-size: 0.9rem; color: #664d03;
}
.simulador-desconto {
    background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
    border: 2px solid #ed1c24; border-radius: 10px;
    padding: 20px; margin: 10px 0;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════
# DADOS PRÉ-CADASTRADOS — FUNÇÕES ENGETROL
# ══════════════════════════════════════════════
FUNCOES_PADRAO = [
    {"nome": "Instrumentista", "salario": 7704.00, "encargos_valor": 13096.80, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Encarregado", "salario": 7329.47, "encargos_valor": 12460.10, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Oficial Montador", "salario": 4264.30, "encargos_valor": 7249.32, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Oficial Eletricista", "salario": 4264.30, "encargos_valor": 7249.32, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": True},
    {"nome": "Ajudante", "salario": 2274.71, "encargos_valor": 3867.01, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Meio Oficial", "salario": 3248.01, "encargos_valor": 5521.61, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Oficial Montador (Alojado)", "salario": 4264.30, "encargos_valor": 7249.32, "beneficio": 1500.00, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Oficial Eletricista (Alojado)", "salario": 4264.30, "encargos_valor": 7249.32, "beneficio": 1500.00, "tipo": "Mão de Obra", "periculosidade": True},
    {"nome": "Supervisor de Obras", "salario": 7384.08, "encargos_valor": 12552.94, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Técnico de Segurança", "salario": 7232.22, "encargos_valor": 12294.77, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
    {"nome": "Administrativo", "salario": 4264.30, "encargos_valor": 7249.32, "beneficio": 0.0, "tipo": "Mão de Obra", "periculosidade": False},
]

ENCARGOS_DETALHADOS_PADRAO = {
    "inss_patronal": 20.00, "sat_rat": 3.00, "salario_educacao": 2.50,
    "sistema_s": 3.30, "fgts": 8.00, "multa_fgts_rescisao": 3.20,
    "decimo_terceiro": 8.33, "ferias_terco": 11.11,
}

# ══════════════════════════════════════════════
# INICIALIZAÇÃO DO SESSION STATE
# ══════════════════════════════════════════════
def _defaults():
    defaults = {
        "nome_obra": "", "cliente": "", "local_obra": "",
        "data_inicio_obra": date.today().isoformat(),
        "lucro_pct": 10.0,
        "modo_lucro": "Lucro sobre PV (BDI)",
        "pis_pct": 1.65, "cofins_pct": 7.60, "iss_pct": 5.00,
        "enc_inss_patronal": ENCARGOS_DETALHADOS_PADRAO["inss_patronal"],
        "enc_sat_rat": ENCARGOS_DETALHADOS_PADRAO["sat_rat"],
        "enc_salario_educacao": ENCARGOS_DETALHADOS_PADRAO["salario_educacao"],
        "enc_sistema_s": ENCARGOS_DETALHADOS_PADRAO["sistema_s"],
        "enc_fgts": ENCARGOS_DETALHADOS_PADRAO["fgts"],
        "enc_multa_fgts": ENCARGOS_DETALHADOS_PADRAO["multa_fgts_rescisao"],
        "enc_decimo_terceiro": ENCARGOS_DETALHADOS_PADRAO["decimo_terceiro"],
        "enc_ferias_terco": ENCARGOS_DETALHADOS_PADRAO["ferias_terco"],
        "divisor_horas": 220.0, "horas_dia": 8.0,
        "pct_horas_extras_50": 0.0, "pct_horas_extras_100": 0.0,
        "markup_risco": 0.0,
        "recursos": [], "recursos_inicializados": False,
        "equipamentos": [],
        "materiais": [],
        "etapas": [],
        "refeicao_valor": 45.0, "hospedagem_valor": 150.0,
        "mobilizacao_valor": 5000.0, "desmobilizacao_valor": 5000.0,
        "diaria_veiculo_valor": 250.0, "qtd_veiculos": 1,
        "distancia_km": 100.0, "consumo_km_litro": 8.0,
        "preco_combustivel": 6.50, "tem_pedagio": False,
        "valor_pedagio_trecho": 0.0, "qtd_viagens_mob": 2,
        "tipo_hospedagem": "Hotel",
        "credito_combustivel": True,
        "credito_locacao_equip": True,
        "orcamento_id": None,
        "desconto_comercial_pct": 0.0,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    if not st.session_state.recursos_inicializados:
        for f in FUNCOES_PADRAO:
            st.session_state.recursos.append({
                "id": str(uuid.uuid4())[:8],
                "nome": f["nome"], "salario": f["salario"],
                "encargos_valor": f["encargos_valor"],
                "beneficio": f["beneficio"], "tipo": f["tipo"],
                "periculosidade": f.get("periculosidade", False),
                "custo_hora": 0.0,
            })
        st.session_state.recursos_inicializados = True

_defaults()


# ══════════════════════════════════════════════
# FUNÇÕES DE CÁLCULO
# ══════════════════════════════════════════════

def total_encargos_pct() -> float:
    ga = (st.session_state.enc_inss_patronal + st.session_state.enc_sat_rat +
          st.session_state.enc_salario_educacao + st.session_state.enc_sistema_s +
          st.session_state.enc_fgts + st.session_state.enc_multa_fgts)
    gb = (st.session_state.enc_decimo_terceiro + st.session_state.enc_ferias_terco)
    gc = ga * gb / 100.0
    return ga + gb + gc


def calcular_custo_hora_com_encargos_fixos(rec: dict) -> float:
    div = st.session_state.divisor_horas
    if div == 0:
        return 0.0
    salario_base = rec["salario"]
    adicional_peric = salario_base * 0.30 if rec.get("periculosidade", False) else 0.0
    custo_mensal = salario_base + adicional_peric + rec["encargos_valor"] + rec.get("beneficio", 0.0)
    return custo_mensal / div


def recalcular_custos_hora():
    for r in st.session_state.recursos:
        r["custo_hora"] = calcular_custo_hora_com_encargos_fixos(r)


def fator_horas_extras() -> float:
    pct50 = st.session_state.pct_horas_extras_50 / 100.0
    pct100 = st.session_state.pct_horas_extras_100 / 100.0
    return 1.0 + (pct50 * 0.50) + (pct100 * 1.00)


def custo_etapa(etapa: dict) -> float:
    recursos_map = {r["id"]: r for r in st.session_state.recursos}
    equip_map = {e["id"]: e for e in st.session_state.equipamentos}
    fhe = fator_horas_extras()
    total = 0.0
    for aloc in etapa.get("alocacoes", []):
        rec = recursos_map.get(aloc["recurso_id"])
        if rec:
            total += etapa["dias"] * st.session_state.horas_dia * aloc["qtd"] * rec["custo_hora"] * fhe
    for aloc in etapa.get("alocacoes_equip", []):
        eq = equip_map.get(aloc["equip_id"])
        if eq:
            total += etapa["dias"] * aloc["qtd"] * eq["custo_diario"]
    total += custo_materiais_etapa(etapa)
    return total


def custo_mo_etapa(etapa: dict) -> float:
    recursos_map = {r["id"]: r for r in st.session_state.recursos}
    fhe = fator_horas_extras()
    total = 0.0
    for aloc in etapa.get("alocacoes", []):
        rec = recursos_map.get(aloc["recurso_id"])
        if rec:
            total += etapa["dias"] * st.session_state.horas_dia * aloc["qtd"] * rec["custo_hora"] * fhe
    return total


def custo_equip_etapa(etapa: dict) -> float:
    equip_map = {e["id"]: e for e in st.session_state.equipamentos}
    total = 0.0
    for aloc in etapa.get("alocacoes_equip", []):
        eq = equip_map.get(aloc["equip_id"])
        if eq:
            total += etapa["dias"] * aloc["qtd"] * eq["custo_diario"]
    return total


def custo_materiais_etapa(etapa: dict) -> float:
    mat_map = {m["id"]: m for m in st.session_state.materiais}
    total = 0.0
    for aloc in etapa.get("alocacoes_mat", []):
        mat = mat_map.get(aloc["material_id"])
        if mat:
            perda = 1.0 + mat.get("margem_perda", 0.0) / 100.0
            total += aloc["qtd"] * perda * mat["preco_unit"]
    return total


def credito_pis_cofins_materiais() -> float:
    aliquota = (st.session_state.pis_pct + st.session_state.cofins_pct) / 100.0
    total_credito = 0.0
    mat_map = {m["id"]: m for m in st.session_state.materiais}
    for et in st.session_state.etapas:
        for aloc in et.get("alocacoes_mat", []):
            mat = mat_map.get(aloc["material_id"])
            if mat and mat.get("recupera_credito", False):
                perda = 1.0 + mat.get("margem_perda", 0.0) / 100.0
                custo_bruto = aloc["qtd"] * perda * mat["preco_unit"]
                total_credito += custo_bruto * aliquota
    return total_credito


def credito_pis_cofins_equip() -> float:
    if not st.session_state.credito_locacao_equip:
        return 0.0
    aliquota = (st.session_state.pis_pct + st.session_state.cofins_pct) / 100.0
    return custo_direto_equip() * aliquota


def credito_pis_cofins_combustivel() -> float:
    if not st.session_state.credito_combustivel:
        return 0.0
    aliquota = (st.session_state.pis_pct + st.session_state.cofins_pct) / 100.0
    ci = custo_indireto_total()
    return ci["combustivel"] * aliquota


def total_creditos_tributarios() -> float:
    return credito_pis_cofins_materiais() + credito_pis_cofins_equip() + credito_pis_cofins_combustivel()


# ══════════════════════════════════════════════
# HISTOGRAMA — LINHA DO TEMPO
# ══════════════════════════════════════════════

def calcular_timeline() -> list:
    """Calcula a linha do tempo dia a dia com empilhamento de equipe."""
    if not st.session_state.etapas:
        return []
    try:
        data_inicio_obra = date.fromisoformat(st.session_state.data_inicio_obra)
    except (ValueError, TypeError):
        data_inicio_obra = date.today()

    recursos_map = {r["id"]: r for r in st.session_state.recursos}
    timeline_data = []

    for etapa in st.session_state.etapas:
        inicio_etapa = etapa.get("data_inicio")
        if inicio_etapa:
            try:
                dt_inicio = date.fromisoformat(inicio_etapa)
            except (ValueError, TypeError):
                dt_inicio = data_inicio_obra
        else:
            dt_inicio = data_inicio_obra

        for dia_offset in range(etapa["dias"]):
            dt = dt_inicio + timedelta(days=dia_offset)
            for aloc in etapa.get("alocacoes", []):
                rec = recursos_map.get(aloc["recurso_id"])
                if rec:
                    for _ in range(aloc["qtd"]):
                        timeline_data.append({
                            "data": dt,
                            "etapa": etapa["nome"],
                            "funcao": rec["nome"],
                            "recurso_id": aloc["recurso_id"],
                        })
    return timeline_data


def histograma_por_dia() -> dict:
    """Retorna {data: total_pessoas} para cada dia da obra."""
    tl = calcular_timeline()
    hist = {}
    for item in tl:
        d = item["data"]
        hist[d] = hist.get(d, 0) + 1
    return dict(sorted(hist.items()))


def histograma_por_funcao() -> dict:
    """Retorna {data: {funcao: qtd}} para cada dia."""
    tl = calcular_timeline()
    hist = {}
    for item in tl:
        d = item["data"]
        f = item["funcao"]
        if d not in hist:
            hist[d] = {}
        hist[d][f] = hist[d].get(f, 0) + 1
    return dict(sorted(hist.items()))


def total_diarias_homem_histograma() -> float:
    """Calcula DH real a partir do histograma (soma de pessoas por dia)."""
    hist = histograma_por_dia()
    return sum(hist.values())


def total_diarias_homem() -> float:
    """Usa histograma se houver datas, senão calcula simples."""
    dh_hist = total_diarias_homem_histograma()
    if dh_hist > 0:
        return dh_hist
    total = 0.0
    for et in st.session_state.etapas:
        pessoas = sum(a["qtd"] for a in et.get("alocacoes", []))
        total += et["dias"] * pessoas
    return total


def total_dias_obra() -> int:
    """Total de dias da obra considerando timeline (com sobreposição)."""
    hist = histograma_por_dia()
    if hist:
        return len(hist)
    return sum(e["dias"] for e in st.session_state.etapas)


def total_dias_obra_sequencial() -> int:
    """Soma simples de dias de todas as etapas (sem considerar sobreposição)."""
    return sum(e["dias"] for e in st.session_state.etapas)


def pico_efetivo() -> int:
    hist = histograma_por_dia()
    if hist:
        return max(hist.values())
    pico = 0
    for et in st.session_state.etapas:
        p = sum(a["qtd"] for a in et.get("alocacoes", []))
        pico = max(pico, p)
    return pico


def custo_direto_mo() -> float:
    return sum(custo_mo_etapa(e) for e in st.session_state.etapas)


def custo_direto_equip() -> float:
    return sum(custo_equip_etapa(e) for e in st.session_state.etapas)


def custo_direto_materiais_bruto() -> float:
    return sum(custo_materiais_etapa(e) for e in st.session_state.etapas)


def custo_direto_total() -> float:
    return custo_direto_mo() + custo_direto_equip() + custo_direto_materiais_bruto()


def custo_indireto_total() -> dict:
    dh = total_diarias_homem()
    dias = total_dias_obra()
    refeicao = st.session_state.refeicao_valor * dh
    hospedagem = st.session_state.hospedagem_valor * dh if st.session_state.tipo_hospedagem == "Hotel" else 0.0
    mob = st.session_state.mobilizacao_valor + st.session_state.desmobilizacao_valor
    veiculos = st.session_state.diaria_veiculo_valor * st.session_state.qtd_veiculos * dias
    dist = st.session_state.distancia_km
    consumo = st.session_state.consumo_km_litro
    preco_comb = st.session_state.preco_combustivel
    viagens = st.session_state.qtd_viagens_mob
    litros = (dist * 2 * viagens) / consumo if consumo > 0 else 0
    combustivel = litros * preco_comb
    pedagio = 0.0
    if st.session_state.tem_pedagio:
        pedagio = st.session_state.valor_pedagio_trecho * 2 * viagens
    total = refeicao + hospedagem + mob + veiculos + combustivel + pedagio
    return {"refeicao": refeicao, "hospedagem": hospedagem, "mobilizacao": mob,
            "veiculos": veiculos, "combustivel": combustivel, "pedagio": pedagio, "total": total}


def custo_total_bruto() -> float:
    return custo_direto_total() + custo_indireto_total()["total"]


def custo_total_liquido() -> float:
    bruto = custo_total_bruto()
    creditos = total_creditos_tributarios()
    risco = bruto * (st.session_state.markup_risco / 100.0)
    return bruto - creditos + risco


def impostos_faturamento_pct() -> float:
    return st.session_state.pis_pct + st.session_state.cofins_pct + st.session_state.iss_pct


def calcular_lucro_bruto_para_bdi(lucro_liquido_desejado_pct: float) -> float:
    """Gross-up: converte lucro líquido desejado em lucro bruto para BDI.
    Considera ~34% de IRPJ+CSLL sobre o lucro (15% + ~10% + 9%)."""
    fator_ir_csll = 0.76  # 1 - 0.15 - 0.09 = 0.76 (conservador, sem adicional)
    return lucro_liquido_desejado_pct / fator_ir_csll


def preco_venda() -> float:
    imp = impostos_faturamento_pct() / 100.0
    lucro_pct = st.session_state.lucro_pct
    if st.session_state.modo_lucro == "Lucro Líquido Alvo (Gross-up)":
        lucro_pct = calcular_lucro_bruto_para_bdi(lucro_pct)
    luc = lucro_pct / 100.0
    divisor = 1 - imp - luc
    if divisor <= 0:
        return 0.0
    return custo_total_liquido() / divisor


def lucro_bruto_pct_efetivo() -> float:
    """Retorna o % de lucro bruto efetivamente usado no BDI."""
    if st.session_state.modo_lucro == "Lucro Líquido Alvo (Gross-up)":
        return calcular_lucro_bruto_para_bdi(st.session_state.lucro_pct)
    return st.session_state.lucro_pct


def preco_venda_com_desconto(desconto_pct: float) -> float:
    pv = preco_venda()
    return pv * (1 - desconto_pct / 100.0)


def calcular_irpj_csll(lucro_mensal: float, meses: int = 1) -> dict:
    lucro_total = lucro_mensal * meses if meses > 0 else lucro_mensal
    limite_total = 20000.0 * meses
    irpj_base = lucro_total * 0.15
    irpj_adicional = max(0, (lucro_total - limite_total)) * 0.10
    irpj_total = irpj_base + irpj_adicional
    csll = lucro_total * 0.09
    return {"lucro_total": lucro_total, "irpj_base": irpj_base,
            "irpj_adicional": irpj_adicional, "irpj_total": irpj_total,
            "csll": csll, "total": irpj_total + csll}


def simular_desconto(desconto_pct: float) -> dict:
    """Simula o impacto de um desconto comercial no PV."""
    pv_original = preco_venda()
    pv_desc = pv_original * (1 - desconto_pct / 100.0)
    ct_liq = custo_total_liquido()
    imp_fat_pct = impostos_faturamento_pct()
    imp_valor = pv_desc * (imp_fat_pct / 100.0)
    rl = pv_desc - imp_valor
    lucro_valor = rl - ct_liq
    margem_pct = (lucro_valor / pv_desc * 100) if pv_desc > 0 else 0
    meses_obra = max(total_dias_obra() / 30, 1)
    irpj = calcular_irpj_csll(lucro_valor / meses_obra, meses_obra) if lucro_valor > 0 else {"total": 0}
    lucro_liquido = lucro_valor - irpj["total"]
    margem_liq_pct = (lucro_liquido / pv_desc * 100) if pv_desc > 0 else 0
    breakeven = lucro_liquido <= 0
    return {
        "pv_original": pv_original, "pv_desconto": pv_desc,
        "desconto_valor": pv_original - pv_desc,
        "lucro_bruto": lucro_valor, "margem_bruta_pct": margem_pct,
        "irpj_csll": irpj["total"], "lucro_liquido": lucro_liquido,
        "margem_liq_pct": margem_liq_pct, "breakeven": breakeven,
    }


def fmt(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_pct(valor: float) -> str:
    return f"{valor:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def validacoes_consistencia() -> list:
    alertas = []
    pe = pico_efetivo()
    veic = st.session_state.qtd_veiculos
    if pe > 0 and veic > pe:
        alertas.append(f"Logística: {veic} veículos para pico de {pe} pessoas. Possível superdimensionamento.")
    if pe > 0 and veic > 0 and pe / veic > 8:
        alertas.append(f"Logística: {pe} pessoas para {veic} veículo(s). Possível subdimensionamento de transporte.")
    dh = total_diarias_homem()
    if dh > 0 and st.session_state.refeicao_valor == 0:
        alertas.append("Refeição com valor zero para obra com mão de obra alocada.")
    if dh > 0 and st.session_state.tipo_hospedagem == "Hotel" and st.session_state.hospedagem_valor == 0:
        alertas.append("Hospedagem em hotel com diária zero.")
    dias = total_dias_obra()
    if dias > 0 and st.session_state.mobilizacao_valor == 0 and st.session_state.desmobilizacao_valor == 0:
        alertas.append("Mobilização e desmobilização com valor zero.")
    ct = custo_total_bruto()
    ci = custo_indireto_total()["total"]
    if ct > 0 and ci / ct > 0.50:
        alertas.append(f"Custos indiretos representam {ci/ct*100:.0f}% do custo total. Verificar proporcionalidade.")
    return alertas


# ══════════════════════════════════════════════
# SERIALIZAÇÃO PARA SQLITE
# ══════════════════════════════════════════════

CAMPOS_SERIALIZAVEIS = [
    "nome_obra", "cliente", "local_obra", "data_inicio_obra",
    "lucro_pct", "modo_lucro",
    "pis_pct", "cofins_pct", "iss_pct",
    "enc_inss_patronal", "enc_sat_rat", "enc_salario_educacao",
    "enc_sistema_s", "enc_fgts", "enc_multa_fgts",
    "enc_decimo_terceiro", "enc_ferias_terco",
    "divisor_horas", "horas_dia",
    "pct_horas_extras_50", "pct_horas_extras_100", "markup_risco",
    "recursos", "equipamentos", "materiais", "etapas",
    "refeicao_valor", "hospedagem_valor",
    "mobilizacao_valor", "desmobilizacao_valor",
    "diaria_veiculo_valor", "qtd_veiculos",
    "distancia_km", "consumo_km_litro", "preco_combustivel",
    "tem_pedagio", "valor_pedagio_trecho", "qtd_viagens_mob",
    "tipo_hospedagem", "credito_combustivel", "credito_locacao_equip",
    "desconto_comercial_pct",
]


def state_to_dict() -> dict:
    return {k: st.session_state[k] for k in CAMPOS_SERIALIZAVEIS if k in st.session_state}


def dict_to_state(d: dict):
    for k, v in d.items():
        if k in CAMPOS_SERIALIZAVEIS:
            st.session_state[k] = v
    st.session_state.recursos_inicializados = True


# ══════════════════════════════════════════════
# GERAÇÃO DE RELATÓRIO MARKDOWN (PARA IA)
# ══════════════════════════════════════════════

def gerar_relatorio_markdown() -> str:
    recalcular_custos_hora()
    cd_mo = custo_direto_mo()
    cd_eq = custo_direto_equip()
    cd_mat = custo_direto_materiais_bruto()
    cd = custo_direto_total()
    ci = custo_indireto_total()
    creditos = total_creditos_tributarios()
    risco_val = custo_total_bruto() * (st.session_state.markup_risco / 100.0)
    ct_liq = custo_total_liquido()
    pv = preco_venda()
    imp_fat_pct = impostos_faturamento_pct()
    imp_valor = pv * (imp_fat_pct / 100.0)
    lucro_bruto_pct = lucro_bruto_pct_efetivo()
    lucro_valor = pv * (lucro_bruto_pct / 100.0)
    rl = pv - imp_valor
    meses_obra = max(total_dias_obra() / 30, 1)
    irpj = calcular_irpj_csll(lucro_valor / meses_obra, meses_obra) if lucro_valor > 0 else {"irpj_base": 0, "irpj_adicional": 0, "irpj_total": 0, "csll": 0, "total": 0}
    lucro_liquido = lucro_valor - irpj["total"]

    md = f"""# RELATÓRIO DE ORÇAMENTO DE OBRA — ENGETROL ENGENHARIA
**Data de Emissão:** {datetime.now().strftime('%d/%m/%Y %H:%M')}
**Regime Tributário:** Lucro Real
**Versão do Sistema:** 3.0
**Modo de Lucro:** {st.session_state.modo_lucro}

---

## 1. IDENTIFICAÇÃO DA OBRA

| Campo | Valor |
|-------|-------|
| Nome da Obra | {st.session_state.nome_obra or 'Não informado'} |
| Cliente | {st.session_state.cliente or 'Não informado'} |
| Local | {st.session_state.local_obra or 'Não informado'} |
| Data de Início | {st.session_state.data_inicio_obra} |
| Total de Dias (calendário) | {total_dias_obra()} dias |
| Total de Dias (sequencial) | {total_dias_obra_sequencial()} dias |
| Pico de Efetivo | {pico_efetivo()} pessoas |
| Total Diárias-Homem | {total_diarias_homem():,.0f} |
| Markup de Risco | {st.session_state.markup_risco:.2f}% |

---

## 2. PREMISSAS TRIBUTÁRIAS (LUCRO REAL)

### 2.1 Impostos sobre Faturamento
| Tributo | Alíquota |
|---------|----------|
| PIS (não-cumulativo) | {st.session_state.pis_pct:.2f}% |
| COFINS (não-cumulativo) | {st.session_state.cofins_pct:.2f}% |
| ISS | {st.session_state.iss_pct:.2f}% |
| **Total sobre Faturamento** | **{imp_fat_pct:.2f}%** |

### 2.2 Impostos sobre o Lucro
| Tributo | Regra | Valor Calculado |
|---------|-------|-----------------|
| IRPJ Base | 15% sobre lucro | {fmt(irpj['irpj_base'])} |
| IRPJ Adicional | 10% sobre excedente de R$ 20.000/mês | {fmt(irpj['irpj_adicional'])} |
| CSLL | 9% sobre lucro | {fmt(irpj['csll'])} |
| **Total IRPJ + CSLL** | | **{fmt(irpj['total'])}** |

### 2.3 Créditos Tributários (PIS/COFINS sobre Insumos)
| Origem do Crédito | Valor |
|-------------------|-------|
| Materiais/Insumos | {fmt(credito_pis_cofins_materiais())} |
| Locação de Equipamentos | {fmt(credito_pis_cofins_equip())} |
| Combustível | {fmt(credito_pis_cofins_combustivel())} |
| **Total Créditos** | **{fmt(creditos)}** |

---

## 3. PREMISSAS DE ENCARGOS SOCIAIS

| Item | % |
|------|---|
| INSS Patronal | {st.session_state.enc_inss_patronal:.2f}% |
| SAT/RAT | {st.session_state.enc_sat_rat:.2f}% |
| Salário Educação | {st.session_state.enc_salario_educacao:.2f}% |
| Sistema S | {st.session_state.enc_sistema_s:.2f}% |
| FGTS | {st.session_state.enc_fgts:.2f}% |
| Provisão Multa FGTS | {st.session_state.enc_multa_fgts:.2f}% |
| 13º Salário | {st.session_state.enc_decimo_terceiro:.2f}% |
| Férias + 1/3 | {st.session_state.enc_ferias_terco:.2f}% |
| **Total Encargos** | **{total_encargos_pct():.2f}%** |

| Parâmetro | Valor |
|-----------|-------|
| Divisor Horas | {st.session_state.divisor_horas:.0f}h |
| Horas/Dia | {st.session_state.horas_dia:.0f}h |
| Horas Extras 50% | {st.session_state.pct_horas_extras_50:.1f}% |
| Horas Extras 100% | {st.session_state.pct_horas_extras_100:.1f}% |
| Fator HE | {fator_horas_extras():.4f} |

---

## 4. BANCO DE RECURSOS (MÃO DE OBRA)

| # | Função | Periculosidade | Salário | Encargos | Benefício | Custo/Hora |
|---|--------|----------------|---------|----------|-----------|------------|
"""
    for i, r in enumerate(st.session_state.recursos, 1):
        peric = "Sim (30%)" if r.get("periculosidade") else "Não"
        md += f"| {i} | {r['nome']} | {peric} | {fmt(r['salario'])} | {fmt(r['encargos_valor'])} | {fmt(r.get('beneficio', 0))} | {fmt(r['custo_hora'])} |\n"

    if st.session_state.equipamentos:
        md += "\n## 5. EQUIPAMENTOS\n\n| # | Equipamento | Custo Diário | Crédito PIS/COFINS |\n|---|-------------|-------------|--------------------|\n"
        for i, e in enumerate(st.session_state.equipamentos, 1):
            md += f"| {i} | {e['nome']} | {fmt(e['custo_diario'])} | {'Sim' if st.session_state.credito_locacao_equip else 'Não'} |\n"

    if st.session_state.materiais:
        md += "\n## 6. MATERIAIS E INSUMOS\n\n| # | Material | Unidade | Preço Unit. | Perda (%) | Recupera Crédito |\n|---|----------|---------|------------|-----------|------------------|\n"
        for i, m in enumerate(st.session_state.materiais, 1):
            md += f"| {i} | {m['nome']} | {m['unidade']} | {fmt(m['preco_unit'])} | {m.get('margem_perda', 0):.1f}% | {'Sim' if m.get('recupera_credito') else 'Não'} |\n"

    md += f"\n---\n\n## 7. CRONOGRAMA E ALOCAÇÃO\n\n"
    recursos_map = {r["id"]: r for r in st.session_state.recursos}

    for i, et in enumerate(st.session_state.etapas, 1):
        data_ini = et.get("data_inicio", "N/I")
        md += f"### Etapa {i}: {et['nome'] or '(sem nome)'}\n"
        md += f"- **Data Início:** {data_ini}\n"
        md += f"- **Duração:** {et['dias']} dias"
        if et.get("modo_hh"):
            md += f" (calculado por HH: {et.get('qtd_servico', 0)} {et.get('unidade_servico', 'un')} x {et.get('indice_hh', 0)} HH/{et.get('unidade_servico', 'un')})"
        md += f"\n- **Custo MO:** {fmt(custo_mo_etapa(et))}\n"
        md += f"- **Custo Equip:** {fmt(custo_equip_etapa(et))}\n"
        md += f"- **Custo Materiais:** {fmt(custo_materiais_etapa(et))}\n"
        md += f"- **Custo Total:** {fmt(custo_etapa(et))}\n\n"

    # Histograma resumo
    hist = histograma_por_dia()
    if hist:
        md += "### Histograma de Efetivo (Resumo)\n\n"
        md += f"- Dias com equipe em campo: {len(hist)}\n"
        md += f"- Pico de efetivo: {max(hist.values())} pessoas\n"
        md += f"- Média de efetivo: {sum(hist.values())/len(hist):.1f} pessoas/dia\n"
        md += f"- Total Diárias-Homem (histograma): {sum(hist.values()):,}\n\n"

    md += f"""---

## 8. CUSTOS INDIRETOS (LOGÍSTICA)

| Item | Valor |
|------|-------|
| Refeição | {fmt(ci['refeicao'])} |
| Hospedagem ({st.session_state.tipo_hospedagem}) | {fmt(ci['hospedagem'])} |
| Mobilização + Desmobilização | {fmt(ci['mobilizacao'])} |
| Veículos | {fmt(ci['veiculos'])} |
| Combustível | {fmt(ci['combustivel'])} |
| Pedágio | {fmt(ci['pedagio'])} |
| **TOTAL** | **{fmt(ci['total'])}** |

---

## 9. DRE GERENCIAL

| Descrição | Valor (R$) | % s/ PV |
|-----------|-----------|---------|
| Receita Bruta (PV) | {fmt(pv)} | 100,00% |
| (-) PIS ({st.session_state.pis_pct:.2f}%) | -{fmt(pv * st.session_state.pis_pct / 100)} | -{st.session_state.pis_pct:.2f}% |
| (-) COFINS ({st.session_state.cofins_pct:.2f}%) | -{fmt(pv * st.session_state.cofins_pct / 100)} | -{st.session_state.cofins_pct:.2f}% |
| (-) ISS ({st.session_state.iss_pct:.2f}%) | -{fmt(pv * st.session_state.iss_pct / 100)} | -{st.session_state.iss_pct:.2f}% |
| (=) Receita Líquida | {fmt(rl)} | {rl/pv*100 if pv > 0 else 0:.2f}% |
| (-) Custo Direto — MO | -{fmt(cd_mo)} | -{cd_mo/pv*100 if pv > 0 else 0:.2f}% |
| (-) Custo Direto — Equipamentos | -{fmt(cd_eq)} | -{cd_eq/pv*100 if pv > 0 else 0:.2f}% |
| (-) Custo Direto — Materiais | -{fmt(cd_mat)} | -{cd_mat/pv*100 if pv > 0 else 0:.2f}% |
| (+) Créditos PIS/COFINS | +{fmt(creditos)} | +{creditos/pv*100 if pv > 0 else 0:.2f}% |
| (-) Custos Indiretos | -{fmt(ci['total'])} | -{ci['total']/pv*100 if pv > 0 else 0:.2f}% |
| (-) Markup de Risco | -{fmt(risco_val)} | -{risco_val/pv*100 if pv > 0 else 0:.2f}% |
| (=) Lucro antes IR/CSLL | {fmt(lucro_valor)} | {lucro_bruto_pct:.2f}% |
| (-) IRPJ | -{fmt(irpj['irpj_total'])} | |
| (-) CSLL | -{fmt(irpj['csll'])} | |
| (=) **LUCRO LÍQUIDO** | **{fmt(lucro_liquido)}** | **{lucro_liquido/pv*100 if pv > 0 else 0:.2f}%** |

---

## 10. FORMAÇÃO DO PREÇO DE VENDA (BDI REVERSO)

| Componente | Valor |
|------------|-------|
| Custo Direto (MO) | {fmt(cd_mo)} |
| Custo Direto (Equip.) | {fmt(cd_eq)} |
| Custo Direto (Materiais) | {fmt(cd_mat)} |
| Custos Indiretos | {fmt(ci['total'])} |
| Custo Bruto Total | {fmt(custo_total_bruto())} |
| (-) Créditos Tributários | -{fmt(creditos)} |
| (+) Markup de Risco ({st.session_state.markup_risco:.1f}%) | +{fmt(risco_val)} |
| **Custo Total Líquido** | **{fmt(ct_liq)}** |
| Impostos s/ Faturamento | {imp_fat_pct:.2f}% |
| Lucro Bruto (BDI) | {lucro_bruto_pct:.2f}% |
| Modo de Lucro | {st.session_state.modo_lucro} |
| Divisor BDI | {1 - imp_fat_pct/100 - lucro_bruto_pct/100:.4f} |
| **PREÇO DE VENDA** | **{fmt(pv)}** |
| BDI Markup | {((pv - ct_liq) / ct_liq * 100) if ct_liq > 0 else 0:.2f}% |

---

## 11. RESUMO EXECUTIVO

| Indicador | Valor |
|-----------|-------|
| Preço de Venda | {fmt(pv)} |
| Custo Total Líquido | {fmt(ct_liq)} |
| Margem Bruta (s/ PV) | {lucro_bruto_pct:.2f}% |
| Margem Líquida (após IRPJ/CSLL) | {lucro_liquido/pv*100 if pv > 0 else 0:.2f}% |
| BDI Markup | {((pv - ct_liq) / ct_liq * 100) if ct_liq > 0 else 0:.2f}% |
| Créditos Tributários Recuperados | {fmt(creditos)} |

---

"""
    alertas = validacoes_consistencia()
    if alertas:
        md += "## 12. ALERTAS DE CONSISTÊNCIA\n\n"
        for a in alertas:
            md += f"- **ALERTA:** {a}\n"
        md += "\n---\n\n"

    md += """*Relatório gerado automaticamente pelo Sistema de Orçamentação v3.0 — Engetrol Engenharia*
*Este documento pode ser compartilhado com IA (ChatGPT/Gemini) para análise de gaps, insights e otimizações.*
"""
    return md


# ══════════════════════════════════════════════
# GERAÇÃO DE PDF
# ══════════════════════════════════════════════

def gerar_pdf_bytes(md_content: str) -> bytes:
    import markdown as md_lib
    html_body = md_lib.markdown(md_content, extensions=["tables", "fenced_code"])
    logo_tag = ""
    if LOGO_B64:
        logo_tag = f'<img src="data:image/jpeg;base64,{LOGO_B64}" style="height:60px; margin-bottom:10px;">'
    html_full = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
@page {{ size: A4; margin: 2cm; @top-center {{ content: "Engetrol Engenharia — Orçamento de Obra"; font-size: 9px; color: #888; }} @bottom-center {{ content: "Página " counter(page) " de " counter(pages); font-size: 9px; color: #888; }} }}
body {{ font-family: 'Helvetica', 'Arial', sans-serif; font-size: 11px; color: #1a1a1a; line-height: 1.5; }}
h1 {{ color: #ed1c24; font-size: 20px; border-bottom: 3px solid #ed1c24; padding-bottom: 8px; }}
h2 {{ color: #1a1a1a; font-size: 15px; border-bottom: 2px solid #ed1c24; padding-bottom: 4px; margin-top: 20px; }}
h3 {{ color: #333; font-size: 13px; }}
table {{ border-collapse: collapse; width: 100%; margin: 10px 0; font-size: 10px; }}
th {{ background: #ed1c24; color: white; padding: 6px 8px; text-align: left; font-weight: 600; }}
td {{ padding: 5px 8px; border-bottom: 1px solid #e0e0e0; }}
tr:nth-child(even) {{ background: #f9f9f9; }}
strong {{ color: #1a1a1a; }}
hr {{ border: none; border-top: 1px solid #ccc; margin: 15px 0; }}
</style></head><body>
<div style="text-align:center; margin-bottom: 20px;">{logo_tag}</div>
{html_body}
</body></html>"""
    from weasyprint import HTML
    return HTML(string=html_full).write_pdf()


# ══════════════════════════════════════════════
# GERAÇÃO DE EXCEL
# ══════════════════════════════════════════════

def gerar_excel_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    red_fill = PatternFill(start_color="ED1C24", end_color="ED1C24", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = red_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def write_row(ws, row, values, bold=False):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            if bold:
                cell.font = bold_font
            cell.border = thin_border

    # ── ABA: Resumo ──
    ws = wb.active
    ws.title = "Resumo"
    ws.cell(row=1, column=1, value="ORÇAMENTO DE OBRA — ENGETROL ENGENHARIA").font = Font(bold=True, size=14, color="ED1C24")
    ws.cell(row=2, column=1, value=f"Obra: {st.session_state.nome_obra or 'N/I'}")
    ws.cell(row=3, column=1, value=f"Cliente: {st.session_state.cliente or 'N/I'}")
    ws.cell(row=4, column=1, value=f"Data: {datetime.now().strftime('%d/%m/%Y')}")

    write_header(ws, 6, ["Indicador", "Valor"])
    resumo_data = [
        ("Custo Direto — MO", custo_direto_mo()),
        ("Custo Direto — Equipamentos", custo_direto_equip()),
        ("Custo Direto — Materiais", custo_direto_materiais_bruto()),
        ("Custos Indiretos", custo_indireto_total()["total"]),
        ("Custo Bruto Total", custo_total_bruto()),
        ("(-) Créditos PIS/COFINS", -total_creditos_tributarios()),
        ("(+) Markup de Risco", custo_total_bruto() * st.session_state.markup_risco / 100),
        ("Custo Total Líquido", custo_total_liquido()),
        ("Preço de Venda", preco_venda()),
    ]
    for i, (ind, val) in enumerate(resumo_data, 7):
        write_row(ws, i, [ind, val], bold=(i == 7 + len(resumo_data) - 1))
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # ── ABA: Mão de Obra ──
    ws_mo = wb.create_sheet("Mão de Obra")
    write_header(ws_mo, 1, ["Função", "Periculosidade", "Salário", "Encargos", "Benefício", "Custo/Hora"])
    for i, r in enumerate(st.session_state.recursos, 2):
        write_row(ws_mo, i, [
            r["nome"], "Sim" if r.get("periculosidade") else "Não",
            r["salario"], r["encargos_valor"], r.get("beneficio", 0), r["custo_hora"]
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_mo.column_dimensions[col].width = 22

    # ── ABA: Materiais ──
    if st.session_state.materiais:
        ws_mat = wb.create_sheet("Materiais")
        write_header(ws_mat, 1, ["Material", "Unidade", "Preço Unit.", "Perda (%)", "Recupera Crédito"])
        for i, m in enumerate(st.session_state.materiais, 2):
            write_row(ws_mat, i, [
                m["nome"], m["unidade"], m["preco_unit"],
                m.get("margem_perda", 0), "Sim" if m.get("recupera_credito") else "Não"
            ])
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_mat.column_dimensions[col].width = 22

    # ── ABA: Cronograma ──
    ws_cron = wb.create_sheet("Cronograma")
    write_header(ws_cron, 1, ["Etapa", "Atividade", "Data Início", "Dias", "Efetivo", "DH", "Custo MO", "Custo Equip", "Custo Mat", "Total"])
    for i, et in enumerate(st.session_state.etapas, 2):
        pessoas = sum(a["qtd"] for a in et.get("alocacoes", []))
        write_row(ws_cron, i, [
            i - 1, et["nome"], et.get("data_inicio", "N/I"), et["dias"], pessoas, et["dias"] * pessoas,
            custo_mo_etapa(et), custo_equip_etapa(et), custo_materiais_etapa(et), custo_etapa(et)
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws_cron.column_dimensions[col].width = 18

    # ── ABA: DRE ──
    ws_dre = wb.create_sheet("DRE")
    pv = preco_venda()
    ci = custo_indireto_total()
    creditos = total_creditos_tributarios()
    risco_val = custo_total_bruto() * st.session_state.markup_risco / 100
    lucro_bruto_p = lucro_bruto_pct_efetivo()
    lucro_valor = pv * lucro_bruto_p / 100
    meses_obra = max(total_dias_obra() / 30, 1)
    irpj = calcular_irpj_csll(lucro_valor / meses_obra, meses_obra) if lucro_valor > 0 else {"irpj_total": 0, "csll": 0, "total": 0}
    lucro_liquido = lucro_valor - irpj["total"]

    write_header(ws_dre, 1, ["Descrição", "Valor (R$)", "% s/ PV"])
    dre_rows = [
        ("Receita Bruta (PV)", pv, "100,00%"),
        (f"(-) PIS ({st.session_state.pis_pct:.2f}%)", -pv * st.session_state.pis_pct / 100, f"-{st.session_state.pis_pct:.2f}%"),
        (f"(-) COFINS ({st.session_state.cofins_pct:.2f}%)", -pv * st.session_state.cofins_pct / 100, f"-{st.session_state.cofins_pct:.2f}%"),
        (f"(-) ISS ({st.session_state.iss_pct:.2f}%)", -pv * st.session_state.iss_pct / 100, f"-{st.session_state.iss_pct:.2f}%"),
        ("(=) Receita Líquida", pv * (1 - impostos_faturamento_pct() / 100), f"{100 - impostos_faturamento_pct():.2f}%"),
        ("(-) Custo Direto — MO", -custo_direto_mo(), ""),
        ("(-) Custo Direto — Equipamentos", -custo_direto_equip(), ""),
        ("(-) Custo Direto — Materiais", -custo_direto_materiais_bruto(), ""),
        ("(+) Créditos PIS/COFINS", creditos, ""),
        ("(-) Custos Indiretos", -ci["total"], ""),
        ("(-) Markup de Risco", -risco_val, ""),
        ("(=) Lucro antes IR/CSLL", lucro_valor, f"{lucro_bruto_p:.2f}%"),
        ("(-) IRPJ", -irpj["irpj_total"], ""),
        ("(-) CSLL", -irpj["csll"], ""),
        ("(=) LUCRO LÍQUIDO", lucro_liquido, f"{lucro_liquido/pv*100 if pv > 0 else 0:.2f}%"),
    ]
    for i, (desc, val, pct) in enumerate(dre_rows, 2):
        write_row(ws_dre, i, [desc, val, pct])
    ws_dre.column_dimensions['A'].width = 40
    ws_dre.column_dimensions['B'].width = 20
    ws_dre.column_dimensions['C'].width = 15

    # ── ABA: Histograma ──
    hist = histograma_por_dia()
    if hist:
        ws_hist = wb.create_sheet("Histograma")
        write_header(ws_hist, 1, ["Data", "Efetivo"])
        for i, (d, qtd) in enumerate(hist.items(), 2):
            write_row(ws_hist, i, [d.strftime("%d/%m/%Y"), qtd])
        ws_hist.column_dimensions['A'].width = 15
        ws_hist.column_dimensions['B'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════
# MENU LATERAL
# ══════════════════════════════════════════════

with st.sidebar:
    if LOGO_B64:
        st.markdown(f'<div style="text-align:center; padding: 10px 0;"><img src="data:image/jpeg;base64,{LOGO_B64}" style="max-width:180px; border-radius:10px;"></div>', unsafe_allow_html=True)
    else:
        st.markdown("## Engetrol Engenharia")
    st.caption("Sistema de Orçamentação v3.0 — Lucro Real")
    st.markdown("---")

    pagina = st.radio("Navegação", [
        "1. Configuração Inicial",
        "2. Banco de Recursos",
        "3. Materiais e Insumos",
        "4. Cronograma e Alocação",
        "5. Custos Indiretos",
        "6. Fechamento",
        "7. Banco de Dados",
    ], label_visibility="collapsed")

    st.markdown("---")
    recalcular_custos_hora()
    ct_val = custo_total_liquido()
    pv_val = preco_venda()
    st.markdown("**Resumo Rápido**")
    st.markdown(f"Custo Total: **{fmt(ct_val)}**")
    st.markdown(f"Preço de Venda: **{fmt(pv_val)}**")
    if pv_val > 0:
        ml = lucro_bruto_pct_efetivo()
        st.markdown(f"Margem Bruta: **{fmt_pct(ml)}**")
    cred = total_creditos_tributarios()
    if cred > 0:
        st.markdown(f"Créditos PIS/COFINS: **{fmt(cred)}**")

    alertas = validacoes_consistencia()
    if alertas:
        st.markdown("---")
        st.markdown("**Alertas**")
        for a in alertas:
            st.warning(a, icon="⚠️")


# ══════════════════════════════════════════════
# PÁGINA 1 — CONFIGURAÇÃO INICIAL
# ══════════════════════════════════════════════

if pagina == "1. Configuração Inicial":
    st.markdown("# Configuração Inicial")
    st.markdown("Defina os parâmetros globais do orçamento. Estes valores serão utilizados em todos os cálculos.")
    st.markdown("---")

    st.markdown("## Dados da Obra")
    co1, co2, co3, co4 = st.columns([2, 2, 2, 1.5])
    with co1:
        st.session_state.nome_obra = st.text_input("Nome da Obra", value=st.session_state.nome_obra, placeholder="Ex: Subestação 138kV")
    with co2:
        st.session_state.cliente = st.text_input("Cliente", value=st.session_state.cliente, placeholder="Ex: Raízen")
    with co3:
        st.session_state.local_obra = st.text_input("Local da Obra", value=st.session_state.local_obra, placeholder="Ex: Piracicaba/SP")
    with co4:
        try:
            dt_ini = date.fromisoformat(st.session_state.data_inicio_obra)
        except (ValueError, TypeError):
            dt_ini = date.today()
        nova_data = st.date_input("Data de Início", value=dt_ini)
        st.session_state.data_inicio_obra = nova_data.isoformat()

    st.markdown("---")

    st.markdown("## Impostos sobre Faturamento (Lucro Real)")
    st.caption("Edite individualmente cada tributo que incide sobre o faturamento bruto.")
    ti1, ti2, ti3, ti4 = st.columns(4)
    with ti1:
        st.session_state.pis_pct = st.number_input("PIS (%)", min_value=0.0, max_value=10.0, value=st.session_state.pis_pct, step=0.05, help="Não-cumulativo: 1,65%")
    with ti2:
        st.session_state.cofins_pct = st.number_input("COFINS (%)", min_value=0.0, max_value=15.0, value=st.session_state.cofins_pct, step=0.10, help="Não-cumulativo: 7,60%")
    with ti3:
        st.session_state.iss_pct = st.number_input("ISS (%)", min_value=0.0, max_value=5.0, value=st.session_state.iss_pct, step=0.50, help="Varia por município: 2% a 5%")
    with ti4:
        total_imp = impostos_faturamento_pct()
        st.metric("Total s/ Faturamento", fmt_pct(total_imp))

    st.markdown("---")

    st.markdown("## Margem de Lucro e Risco")
    lu0, lu1, lu2, lu3 = st.columns(4)
    with lu0:
        modo_lucro_opcoes = ["Lucro sobre PV (BDI)", "Lucro Líquido Alvo (Gross-up)"]
        idx_modo = modo_lucro_opcoes.index(st.session_state.modo_lucro) if st.session_state.modo_lucro in modo_lucro_opcoes else 0
        st.session_state.modo_lucro = st.selectbox("Modo de Cálculo do Lucro", modo_lucro_opcoes, index=idx_modo,
            help="BDI: lucro % sobre PV antes de IR. Gross-up: o sistema calcula o lucro bruto necessário para que o lucro LÍQUIDO (após IRPJ/CSLL) seja o % desejado.")
    with lu1:
        label_lucro = "Lucro Líquido Desejado (%)" if st.session_state.modo_lucro == "Lucro Líquido Alvo (Gross-up)" else "Lucro Desejado (%)"
        st.session_state.lucro_pct = st.number_input(label_lucro, min_value=0.0, max_value=50.0, value=st.session_state.lucro_pct, step=0.5)
    with lu2:
        st.session_state.markup_risco = st.number_input("Markup de Risco/Contingência (%)", min_value=0.0, max_value=20.0, value=st.session_state.markup_risco, step=0.5, help="Cobertura para imprevistos técnicos e variação de preços")
    with lu3:
        lucro_bdi = lucro_bruto_pct_efetivo()
        soma = total_imp + lucro_bdi
        if soma >= 100:
            st.error(f"Impostos + Lucro = {soma:.2f}% (deve ser < 100%)")
        else:
            st.info(f"Divisor BDI: {(100 - soma):.2f}%")
            if st.session_state.modo_lucro == "Lucro Líquido Alvo (Gross-up)":
                st.caption(f"Lucro bruto necessário: {lucro_bdi:.2f}% para atingir {st.session_state.lucro_pct:.2f}% líquido")

    st.markdown("---")

    st.markdown("## IRPJ e CSLL (sobre o Lucro)")
    with st.expander("Regras do IRPJ e CSLL — Lucro Real", expanded=False):
        st.markdown("""
| Tributo | Regra | Base Legal |
|---------|-------|------------|
| **IRPJ** | 15% sobre o lucro real | Lei 9.249/95, art. 3º |
| **IRPJ Adicional** | 10% sobre parcela do lucro que exceder R$ 20.000/mês | Lei 9.249/95, art. 3º, §1º |
| **CSLL** | 9% sobre o lucro líquido ajustado | Lei 7.689/88, art. 3º |

> **Importante:** IRPJ e CSLL incidem sobre o **lucro apurado**, não sobre o faturamento.
> O adicional de 10% aplica-se apenas à parcela que exceder R$ 20.000,00/mês.
        """)

    st.markdown("---")

    st.markdown("## Encargos Sociais sobre Salário")
    st.caption("Composição detalhada conforme legislação vigente (CLT, Lei 8.212/91, CF/88).")

    st.markdown("### Grupo A — Encargos Sociais Básicos")
    ea1, ea2, ea3 = st.columns(3)
    with ea1:
        st.session_state.enc_inss_patronal = st.number_input("INSS Patronal (%)", min_value=0.0, max_value=30.0, value=st.session_state.enc_inss_patronal, step=0.5, help="Art. 22, Lei 8.212/91 — 20%")
        st.session_state.enc_sat_rat = st.number_input("SAT/RAT (%)", min_value=0.0, max_value=6.0, value=st.session_state.enc_sat_rat, step=0.5, help="Grau de risco 3: 3%")
    with ea2:
        st.session_state.enc_salario_educacao = st.number_input("Salário Educação (%)", min_value=0.0, max_value=5.0, value=st.session_state.enc_salario_educacao, step=0.1, help="Art. 212, §5º CF — 2,5%")
        st.session_state.enc_sistema_s = st.number_input("Sistema S (%)", min_value=0.0, max_value=6.0, value=st.session_state.enc_sistema_s, step=0.1, help="SESI+SENAI+SEBRAE+INCRA = 3,3%")
    with ea3:
        st.session_state.enc_fgts = st.number_input("FGTS (%)", min_value=0.0, max_value=12.0, value=st.session_state.enc_fgts, step=0.5, help="Art. 15, Lei 8.036/90 — 8%")
        st.session_state.enc_multa_fgts = st.number_input("Provisão Multa FGTS (%)", min_value=0.0, max_value=6.0, value=st.session_state.enc_multa_fgts, step=0.1, help="40% × 8% = 3,2%")

    ga = (st.session_state.enc_inss_patronal + st.session_state.enc_sat_rat +
          st.session_state.enc_salario_educacao + st.session_state.enc_sistema_s +
          st.session_state.enc_fgts + st.session_state.enc_multa_fgts)
    st.markdown(f"**Subtotal Grupo A:** {fmt_pct(ga)}")

    st.markdown("### Grupo B — Provisões Trabalhistas")
    eb1, eb2 = st.columns(2)
    with eb1:
        st.session_state.enc_decimo_terceiro = st.number_input("13º Salário — 1/12 (%)", min_value=0.0, max_value=15.0, value=st.session_state.enc_decimo_terceiro, step=0.01, help="Art. 7º, VIII, CF — 8,33%")
    with eb2:
        st.session_state.enc_ferias_terco = st.number_input("Férias + 1/3 Constitucional (%)", min_value=0.0, max_value=20.0, value=st.session_state.enc_ferias_terco, step=0.01, help="Art. 7º, XVII, CF — 11,11%")

    gb = st.session_state.enc_decimo_terceiro + st.session_state.enc_ferias_terco
    gc = ga * gb / 100.0
    total_enc = ga + gb + gc
    st.markdown(f"**Subtotal Grupo B:** {fmt_pct(gb)}")
    st.markdown(f"**Grupo C (Incidência A sobre B):** {fmt_pct(gc)}")
    st.success(f"**Total de Encargos Sociais: {fmt_pct(total_enc)}**")

    st.markdown("---")

    st.markdown("## Parâmetros de Cálculo de Hora")
    ph1, ph2, ph3, ph4 = st.columns(4)
    with ph1:
        st.session_state.divisor_horas = st.number_input("Divisor de Horas Mensais (h)", min_value=1.0, max_value=300.0, value=st.session_state.divisor_horas, step=8.0, help="CLT: 220h")
    with ph2:
        st.session_state.horas_dia = st.number_input("Horas de Trabalho por Dia (h)", min_value=1.0, max_value=24.0, value=st.session_state.horas_dia, step=1.0)
    with ph3:
        st.session_state.pct_horas_extras_50 = st.number_input("% Horas Extras 50%", min_value=0.0, max_value=100.0, value=st.session_state.pct_horas_extras_50, step=5.0, help="Percentual de HE a 50% sobre o total")
    with ph4:
        st.session_state.pct_horas_extras_100 = st.number_input("% Horas Extras 100%", min_value=0.0, max_value=100.0, value=st.session_state.pct_horas_extras_100, step=5.0, help="Percentual de HE a 100% (domingos/feriados)")

    fhe = fator_horas_extras()
    if fhe > 1.0:
        st.info(f"Fator de Horas Extras: **{fhe:.4f}** (custo de MO será multiplicado por este fator)")

    recalcular_custos_hora()


# ══════════════════════════════════════════════
# PÁGINA 2 — BANCO DE RECURSOS
# ══════════════════════════════════════════════

elif pagina == "2. Banco de Recursos":
    st.markdown("# Banco de Recursos")
    st.markdown("Funções pré-cadastradas da Engetrol e equipamentos. Edite salários, adicione ou remova conforme necessário.")
    st.markdown("---")
    recalcular_custos_hora()

    tab_mo, tab_eq = st.tabs(["Mão de Obra", "Equipamentos"])

    with tab_mo:
        st.markdown("## Funções Cadastradas")
        if st.session_state.recursos:
            for idx, rec in enumerate(st.session_state.recursos):
                peric_label = " | NR-10 (30%)" if rec.get("periculosidade") else ""
                with st.expander(f"{rec['nome']}{peric_label} — Custo/Hora: {fmt(rec['custo_hora'])}", expanded=False):
                    c1, c2, c3, c4, c5 = st.columns([2, 1.5, 1.5, 1.5, 1])
                    with c1:
                        novo_nome = st.text_input("Nome", value=rec["nome"], key=f"rn_{rec['id']}")
                    with c2:
                        novo_sal = st.number_input("Salário Base (R$)", min_value=0.0, value=rec["salario"], step=100.0, key=f"rs_{rec['id']}")
                    with c3:
                        novo_enc = st.number_input("Encargos (R$)", min_value=0.0, value=rec["encargos_valor"], step=100.0, key=f"re_{rec['id']}")
                    with c4:
                        novo_ben = st.number_input("Benefício (R$)", min_value=0.0, value=rec.get("beneficio", 0.0), step=100.0, key=f"rb_{rec['id']}")
                    with c5:
                        novo_peric = st.checkbox("Periculosidade NR-10", value=rec.get("periculosidade", False), key=f"rp_{rec['id']}")

                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("Atualizar", key=f"rupd_{rec['id']}", use_container_width=True):
                            st.session_state.recursos[idx].update({
                                "nome": novo_nome, "salario": novo_sal,
                                "encargos_valor": novo_enc, "beneficio": novo_ben,
                                "periculosidade": novo_peric,
                            })
                            recalcular_custos_hora()
                            st.rerun()
                    with bc2:
                        if st.button("Remover", key=f"rdel_{rec['id']}", type="secondary", use_container_width=True):
                            rid = rec["id"]
                            st.session_state.recursos.pop(idx)
                            for et in st.session_state.etapas:
                                et["alocacoes"] = [a for a in et.get("alocacoes", []) if a["recurso_id"] != rid]
                            st.rerun()

            # Tabela resumo
            df_data = []
            for r in st.session_state.recursos:
                df_data.append({
                    "Função": r["nome"],
                    "Periculosidade": "Sim" if r.get("periculosidade") else "Não",
                    "Salário": f"R$ {r['salario']:,.2f}",
                    "Encargos": f"R$ {r['encargos_valor']:,.2f}",
                    "Benefício": f"R$ {r.get('beneficio', 0):,.2f}",
                    "Custo/Hora": f"R$ {r['custo_hora']:,.2f}",
                })
            st.dataframe(pd.DataFrame(df_data), use_container_width=True, hide_index=True)

        st.markdown("### Cadastrar Nova Função")
        with st.form("form_novo_recurso", clear_on_submit=True):
            nc1, nc2, nc3, nc4, nc5 = st.columns([2, 1.5, 1.5, 1.5, 1])
            with nc1:
                nr_nome = st.text_input("Nome da Função", placeholder="Ex: Soldador")
            with nc2:
                nr_sal = st.number_input("Salário Base (R$)", min_value=0.0, value=3000.0, step=100.0)
            with nc3:
                nr_enc = st.number_input("Encargos (R$)", min_value=0.0, value=5100.0, step=100.0)
            with nc4:
                nr_ben = st.number_input("Benefício (R$)", min_value=0.0, value=0.0, step=100.0)
            with nc5:
                nr_peric = st.checkbox("Periculosidade", value=False)

            if st.form_submit_button("Cadastrar Função", use_container_width=True):
                if nr_nome.strip():
                    novo = {
                        "id": str(uuid.uuid4())[:8], "nome": nr_nome.strip(),
                        "salario": nr_sal, "encargos_valor": nr_enc,
                        "beneficio": nr_ben, "tipo": "Mão de Obra",
                        "periculosidade": nr_peric, "custo_hora": 0.0,
                    }
                    st.session_state.recursos.append(novo)
                    recalcular_custos_hora()
                    st.rerun()
                else:
                    st.warning("Informe o nome da função.")

    with tab_eq:
        st.markdown("## Equipamentos")
        st.session_state.credito_locacao_equip = st.checkbox(
            "Equipamentos geram crédito PIS/COFINS (locação)?",
            value=st.session_state.credito_locacao_equip,
            help="No Lucro Real, locação de máquinas e equipamentos gera crédito de PIS/COFINS (9,25%)"
        )

        if st.session_state.equipamentos:
            for idx, eq in enumerate(st.session_state.equipamentos):
                with st.expander(f"{eq['nome']} — {fmt(eq['custo_diario'])}/dia", expanded=False):
                    eq1, eq2 = st.columns(2)
                    with eq1:
                        en = st.text_input("Nome", value=eq["nome"], key=f"eqn_{eq['id']}")
                    with eq2:
                        ec = st.number_input("Custo Diário (R$)", min_value=0.0, value=eq["custo_diario"], step=50.0, key=f"eqc_{eq['id']}")
                    ebc1, ebc2 = st.columns(2)
                    with ebc1:
                        if st.button("Atualizar", key=f"equpd_{eq['id']}", use_container_width=True):
                            st.session_state.equipamentos[idx]["nome"] = en
                            st.session_state.equipamentos[idx]["custo_diario"] = ec
                            st.rerun()
                    with ebc2:
                        if st.button("Remover", key=f"eqdel_{eq['id']}", type="secondary", use_container_width=True):
                            eid = eq["id"]
                            st.session_state.equipamentos.pop(idx)
                            for et in st.session_state.etapas:
                                et["alocacoes_equip"] = [a for a in et.get("alocacoes_equip", []) if a["equip_id"] != eid]
                            st.rerun()

        st.markdown("### Cadastrar Novo Equipamento")
        with st.form("form_novo_equip", clear_on_submit=True):
            neq1, neq2 = st.columns(2)
            with neq1:
                neq_nome = st.text_input("Nome do Equipamento", placeholder="Ex: Guindaste 30t")
            with neq2:
                neq_custo = st.number_input("Custo Diário (R$)", min_value=0.0, value=500.0, step=50.0)
            if st.form_submit_button("Cadastrar Equipamento", use_container_width=True):
                if neq_nome.strip():
                    st.session_state.equipamentos.append({"id": str(uuid.uuid4())[:8], "nome": neq_nome.strip(), "custo_diario": neq_custo})
                    st.rerun()
                else:
                    st.warning("Informe o nome do equipamento.")


# ══════════════════════════════════════════════
# PÁGINA 3 — MATERIAIS E INSUMOS (BOM)
# ══════════════════════════════════════════════

elif pagina == "3. Materiais e Insumos":
    st.markdown("# Materiais e Insumos (BOM)")
    st.markdown("Cadastre materiais para alocação nas etapas. No Lucro Real, insumos com crédito de PIS/COFINS reduzem o custo efetivo.")
    st.markdown("---")

    # IMPORTAÇÃO DE PLANILHA
    st.markdown("## Importar Lista de Materiais")
    st.caption("Carregue um arquivo Excel (.xlsx) ou CSV com colunas: Descrição, Unidade, Preço (ou Preço Unitário)")
    uploaded_file = st.file_uploader("Importar Materiais (Excel/CSV)", type=["xlsx", "csv"], label_visibility="collapsed")
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                df_imp = pd.read_csv(uploaded_file)
            else:
                df_imp = pd.read_excel(uploaded_file)

            st.markdown("**Prévia dos dados importados:**")
            st.dataframe(df_imp.head(10), use_container_width=True, hide_index=True)

            # Mapeamento de colunas
            colunas = list(df_imp.columns)
            st.markdown("**Mapeamento de colunas:**")
            mc1, mc2, mc3 = st.columns(3)
            with mc1:
                col_desc = st.selectbox("Coluna Descrição", colunas, index=0)
            with mc2:
                col_unid = st.selectbox("Coluna Unidade", ["(nenhuma)"] + colunas, index=0)
            with mc3:
                col_preco = st.selectbox("Coluna Preço", colunas, index=min(1, len(colunas) - 1) if len(colunas) > 1 else 0)

            if st.button("Importar Materiais", use_container_width=True):
                importados = 0
                for _, row in df_imp.iterrows():
                    desc = str(row.get(col_desc, "")).strip()
                    unid = str(row.get(col_unid, "un")).strip() if col_unid != "(nenhuma)" else "un"
                    try:
                        preco = float(str(row.get(col_preco, 0)).replace("R$", "").replace(".", "").replace(",", ".").strip())
                    except (ValueError, TypeError):
                        preco = 0.0
                    if desc and preco > 0:
                        st.session_state.materiais.append({
                            "id": str(uuid.uuid4())[:8], "nome": desc,
                            "unidade": unid if unid in ["m", "un", "kg", "cx", "par", "pç", "L", "rolo"] else "un",
                            "preco_unit": preco, "margem_perda": 5.0, "recupera_credito": True,
                        })
                        importados += 1
                st.success(f"{importados} materiais importados com sucesso!")
                st.rerun()
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")

    st.markdown("---")

    if st.session_state.materiais:
        st.markdown("## Materiais Cadastrados")
        for idx, mat in enumerate(st.session_state.materiais):
            credito_label = " | Crédito PIS/COFINS" if mat.get("recupera_credito") else ""
            with st.expander(f"{mat['nome']} — {fmt(mat['preco_unit'])}/{mat['unidade']}{credito_label}", expanded=False):
                m1, m2, m3, m4, m5 = st.columns([2.5, 1, 1.5, 1, 1])
                with m1:
                    mn = st.text_input("Descrição", value=mat["nome"], key=f"mn_{mat['id']}")
                with m2:
                    mu = st.selectbox("Unidade", ["m", "un", "kg", "cx", "par", "pç", "L", "rolo"],
                                      index=["m", "un", "kg", "cx", "par", "pç", "L", "rolo"].index(mat["unidade"]) if mat["unidade"] in ["m", "un", "kg", "cx", "par", "pç", "L", "rolo"] else 0,
                                      key=f"mu_{mat['id']}")
                with m3:
                    mp = st.number_input("Preço Unit. (R$)", min_value=0.0, value=mat["preco_unit"], step=1.0, key=f"mp_{mat['id']}")
                with m4:
                    mw = st.number_input("Perda (%)", min_value=0.0, max_value=50.0, value=mat.get("margem_perda", 0.0), step=1.0, key=f"mw_{mat['id']}")
                with m5:
                    mc = st.checkbox("Crédito?", value=mat.get("recupera_credito", True), key=f"mc_{mat['id']}")

                mbc1, mbc2 = st.columns(2)
                with mbc1:
                    if st.button("Atualizar", key=f"mupd_{mat['id']}", use_container_width=True):
                        st.session_state.materiais[idx].update({"nome": mn, "unidade": mu, "preco_unit": mp, "margem_perda": mw, "recupera_credito": mc})
                        st.rerun()
                with mbc2:
                    if st.button("Remover", key=f"mdel_{mat['id']}", type="secondary", use_container_width=True):
                        mid = mat["id"]
                        st.session_state.materiais.pop(idx)
                        for et in st.session_state.etapas:
                            et["alocacoes_mat"] = [a for a in et.get("alocacoes_mat", []) if a["material_id"] != mid]
                        st.rerun()

        st.markdown("### Resumo de Materiais")
        df_mat = pd.DataFrame([{
            "Material": m["nome"], "Unidade": m["unidade"],
            "Preço Unit.": f"R$ {m['preco_unit']:,.2f}",
            "Perda (%)": f"{m.get('margem_perda', 0):.1f}%",
            "Crédito": "Sim" if m.get("recupera_credito") else "Não",
        } for m in st.session_state.materiais])
        st.dataframe(df_mat, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("## Cadastrar Novo Material")
    with st.form("form_novo_material", clear_on_submit=True):
        nm1, nm2, nm3, nm4 = st.columns([3, 1, 1.5, 1])
        with nm1:
            mat_nome = st.text_input("Descrição do Material", placeholder="Ex: Cabo Flexível 4mm² - Vermelho")
        with nm2:
            mat_unidade = st.selectbox("Unidade", ["m", "un", "kg", "cx", "par", "pç", "L", "rolo"])
        with nm3:
            mat_preco = st.number_input("Preço Unitário (R$)", min_value=0.0, step=1.0, value=0.0)
        with nm4:
            mat_perda = st.number_input("Perda (%)", min_value=0.0, max_value=50.0, step=1.0, value=5.0, help="Margem de perda técnica")

        mat_credito = st.checkbox("Recupera Crédito PIS/COFINS?", value=True, help="Insumos no Lucro Real geram crédito de 9,25%")

        if st.form_submit_button("Cadastrar Material", use_container_width=True):
            if mat_nome.strip() and mat_preco > 0:
                st.session_state.materiais.append({
                    "id": str(uuid.uuid4())[:8], "nome": mat_nome.strip(),
                    "unidade": mat_unidade, "preco_unit": mat_preco,
                    "margem_perda": mat_perda, "recupera_credito": mat_credito,
                })
                st.rerun()
            else:
                st.warning("Informe a descrição e o preço do material.")

    # Curva ABC
    if st.session_state.materiais and st.session_state.etapas:
        st.markdown("---")
        st.markdown("## Curva ABC de Materiais")
        mat_map = {m["id"]: m for m in st.session_state.materiais}
        custos_mat = {}
        for et in st.session_state.etapas:
            for aloc in et.get("alocacoes_mat", []):
                mat = mat_map.get(aloc["material_id"])
                if mat:
                    perda = 1.0 + mat.get("margem_perda", 0.0) / 100.0
                    custo = aloc["qtd"] * perda * mat["preco_unit"]
                    custos_mat[mat["nome"]] = custos_mat.get(mat["nome"], 0) + custo

        if custos_mat:
            df_abc = pd.DataFrame(list(custos_mat.items()), columns=["Material", "Custo Total"])
            df_abc = df_abc.sort_values("Custo Total", ascending=False)
            df_abc["% Acumulado"] = (df_abc["Custo Total"].cumsum() / df_abc["Custo Total"].sum() * 100)
            df_abc["Classe"] = df_abc["% Acumulado"].apply(lambda x: "A" if x <= 80 else ("B" if x <= 95 else "C"))
            df_abc["Custo Total"] = df_abc["Custo Total"].apply(lambda x: f"R$ {x:,.2f}")
            df_abc["% Acumulado"] = df_abc["% Acumulado"].apply(lambda x: f"{x:.1f}%")
            st.dataframe(df_abc, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════
# PÁGINA 4 — CRONOGRAMA E ALOCAÇÃO
# ══════════════════════════════════════════════

elif pagina == "4. Cronograma e Alocação":
    st.markdown("# Cronograma e Alocação")
    st.markdown("Construa as etapas da obra com datas, recursos, equipamentos e materiais. O histograma de efetivo é gerado automaticamente.")
    st.markdown("---")
    recalcular_custos_hora()

    # Formulário de nova etapa
    st.markdown("## Adicionar Nova Etapa")
    with st.form("form_nova_etapa", clear_on_submit=True):
        et1, et2, et3 = st.columns([3, 1.5, 1.5])
        with et1:
            et_nome = st.text_input("Nome da Atividade", placeholder="Ex: Montagem de Painéis")
        with et2:
            try:
                dt_ini = date.fromisoformat(st.session_state.data_inicio_obra)
            except (ValueError, TypeError):
                dt_ini = date.today()
            et_data_inicio = st.date_input("Data de Início da Etapa", value=dt_ini)
        with et3:
            et_dias = st.number_input("Duração (Dias)", min_value=1, value=5, step=1)

        # Modo HH
        modo_hh = st.checkbox("Calcular por Produtividade (HH)", value=False, help="Informe quantidade de serviço e índice HH/unidade")
        qtd_servico = 0.0
        unidade_servico = "m"
        indice_hh = 0.0
        if modo_hh:
            hh1, hh2, hh3 = st.columns(3)
            with hh1:
                qtd_servico = st.number_input("Quantidade de Serviço", min_value=0.0, value=100.0, step=10.0)
            with hh2:
                unidade_servico = st.selectbox("Unidade de Serviço", ["m", "m²", "m³", "un", "kg", "ponto", "circuito"])
            with hh3:
                indice_hh = st.number_input("Índice (HH por unidade)", min_value=0.0, value=0.5, step=0.1, help="Ex: 0.5 HH/metro de cabo")
            total_hh = qtd_servico * indice_hh
            if total_hh > 0:
                st.info(f"Total HH: **{total_hh:,.1f}** | Duração sugerida: **{math.ceil(total_hh / st.session_state.horas_dia):.0f} dias** (1 pessoa)")

        # Seleção de recursos
        st.markdown("**Mão de Obra:**")
        rec_opcoes = {r["id"]: f"{r['nome']} — {fmt(r['custo_hora'])}/h" for r in st.session_state.recursos}
        sel_recursos = st.multiselect("Selecione os recursos", options=list(rec_opcoes.keys()), format_func=lambda x: rec_opcoes[x])
        qtds = {}
        if sel_recursos:
            cols_r = st.columns(min(len(sel_recursos), 4))
            for i, rid in enumerate(sel_recursos):
                rec = next((r for r in st.session_state.recursos if r["id"] == rid), None)
                if rec:
                    with cols_r[i % len(cols_r)]:
                        qtds[rid] = st.number_input(f"Qtd {rec['nome']}", min_value=1, value=1, step=1, key=f"nq_{rid}")

        # Seleção de equipamentos
        if st.session_state.equipamentos:
            st.markdown("**Equipamentos:**")
            eq_opcoes = {e["id"]: f"{e['nome']} — {fmt(e['custo_diario'])}/dia" for e in st.session_state.equipamentos}
            sel_equips = st.multiselect("Selecione os equipamentos", options=list(eq_opcoes.keys()), format_func=lambda x: eq_opcoes[x])
            qtds_eq = {}
            if sel_equips:
                cols_eq = st.columns(min(len(sel_equips), 4))
                for i, eid in enumerate(sel_equips):
                    eq = next((e for e in st.session_state.equipamentos if e["id"] == eid), None)
                    if eq:
                        with cols_eq[i % len(cols_eq)]:
                            qtds_eq[eid] = st.number_input(f"Qtd {eq['nome']}", min_value=1, value=1, step=1, key=f"nqe_{eid}")
        else:
            sel_equips = []
            qtds_eq = {}

        # Seleção de materiais
        if st.session_state.materiais:
            st.markdown("**Materiais:**")
            mat_opcoes = {m["id"]: f"{m['nome']} ({m['unidade']}) — {fmt(m['preco_unit'])}" for m in st.session_state.materiais}
            sel_mats = st.multiselect("Selecione os materiais", options=list(mat_opcoes.keys()), format_func=lambda x: mat_opcoes[x])
            qtds_mat = {}
            if sel_mats:
                cols_mat = st.columns(min(len(sel_mats), 4))
                for i, mid in enumerate(sel_mats):
                    mat = next((m for m in st.session_state.materiais if m["id"] == mid), None)
                    if mat:
                        with cols_mat[i % len(cols_mat)]:
                            qtds_mat[mid] = st.number_input(f"Qtd {mat['nome']} ({mat['unidade']})", min_value=0.1, value=1.0, step=1.0, key=f"nqm_{mid}")
        else:
            sel_mats = []
            qtds_mat = {}

        submitted = st.form_submit_button("Confirmar Etapa", use_container_width=True)
        if submitted:
            if et_nome.strip() and (sel_recursos or sel_equips or sel_mats):
                alocacoes = [{"recurso_id": rid, "qtd": qtds.get(rid, 1)} for rid in sel_recursos]
                alocacoes_eq = [{"equip_id": eid, "qtd": qtds_eq.get(eid, 1)} for eid in sel_equips]
                alocacoes_mat = [{"material_id": mid, "qtd": qtds_mat.get(mid, 1.0)} for mid in sel_mats]
                nova_etapa = {
                    "id": str(uuid.uuid4())[:8], "nome": et_nome.strip(),
                    "data_inicio": et_data_inicio.isoformat(),
                    "dias": et_dias, "alocacoes": alocacoes,
                    "alocacoes_equip": alocacoes_eq, "alocacoes_mat": alocacoes_mat,
                    "modo_hh": modo_hh, "qtd_servico": qtd_servico,
                    "unidade_servico": unidade_servico, "indice_hh": indice_hh,
                }
                st.session_state.etapas.append(nova_etapa)
                st.success(f"Etapa **{et_nome}** adicionada!")
                st.rerun()
            elif not et_nome.strip():
                st.warning("Informe o nome da atividade.")
            else:
                st.warning("Selecione ao menos um recurso, equipamento ou material.")

    st.markdown("---")
    st.markdown("## Etapas Cadastradas")

    if not st.session_state.etapas:
        st.info("Nenhuma etapa criada. Preencha o formulário acima e clique em **Confirmar Etapa**.")
    else:
        recursos_map = {r["id"]: r for r in st.session_state.recursos}
        equip_map = {e["id"]: e for e in st.session_state.equipamentos}
        mat_map = {m["id"]: m for m in st.session_state.materiais}
        custo_acum = 0.0

        for idx, etapa in enumerate(st.session_state.etapas):
            ce = custo_etapa(etapa)
            custo_acum += ce
            pessoas = sum(a["qtd"] for a in etapa.get("alocacoes", []))
            hh_label = ""
            if etapa.get("modo_hh"):
                hh_label = f" | HH: {etapa.get('qtd_servico', 0)} {etapa.get('unidade_servico', '')} × {etapa.get('indice_hh', 0)} HH"

            with st.expander(f"Etapa {idx+1}: {etapa['nome']} — {etapa.get('data_inicio', 'N/I')} — {etapa['dias']}d — {pessoas} pess. — {fmt(ce)}{hh_label}", expanded=False):
                if etapa.get("alocacoes"):
                    det = []
                    fhe = fator_horas_extras()
                    for a in etapa["alocacoes"]:
                        rec = recursos_map.get(a["recurso_id"])
                        if rec:
                            sub = etapa["dias"] * st.session_state.horas_dia * a["qtd"] * rec["custo_hora"] * fhe
                            det.append({"Função": rec["nome"], "Qtd": a["qtd"], "Custo/Hora": fmt(rec["custo_hora"]), "Subtotal": fmt(sub)})
                    st.dataframe(pd.DataFrame(det), use_container_width=True, hide_index=True)

                if etapa.get("alocacoes_equip"):
                    det_eq = []
                    for a in etapa["alocacoes_equip"]:
                        eq = equip_map.get(a["equip_id"])
                        if eq:
                            sub = etapa["dias"] * a["qtd"] * eq["custo_diario"]
                            det_eq.append({"Equipamento": eq["nome"], "Qtd": a["qtd"], "Custo/Dia": fmt(eq["custo_diario"]), "Subtotal": fmt(sub)})
                    st.dataframe(pd.DataFrame(det_eq), use_container_width=True, hide_index=True)

                if etapa.get("alocacoes_mat"):
                    det_mat = []
                    for a in etapa["alocacoes_mat"]:
                        mat = mat_map.get(a["material_id"])
                        if mat:
                            perda = 1.0 + mat.get("margem_perda", 0.0) / 100.0
                            sub = a["qtd"] * perda * mat["preco_unit"]
                            det_mat.append({"Material": mat["nome"], "Qtd": f"{a['qtd']:,.1f} {mat['unidade']}", "Preço": fmt(mat["preco_unit"]), "Perda": f"{mat.get('margem_perda', 0):.0f}%", "Subtotal": fmt(sub)})
                    st.dataframe(pd.DataFrame(det_mat), use_container_width=True, hide_index=True)

                if st.button("Remover Etapa", key=f"del_et_{etapa['id']}", type="secondary"):
                    st.session_state.etapas.pop(idx)
                    st.rerun()

        st.markdown("---")
        st.markdown(f"### Custo Direto Total: **{fmt(custo_acum)}**")

        resumo = []
        for idx, et in enumerate(st.session_state.etapas):
            pessoas = sum(a["qtd"] for a in et.get("alocacoes", []))
            resumo.append({
                "Etapa": f"{idx+1}. {et['nome']}",
                "Data Início": et.get("data_inicio", "N/I"),
                "Dias": et["dias"], "Efetivo": pessoas,
                "DH": et["dias"] * pessoas,
                "Custo MO": fmt(custo_mo_etapa(et)),
                "Custo Equip": fmt(custo_equip_etapa(et)),
                "Custo Mat": fmt(custo_materiais_etapa(et)),
                "Total": fmt(custo_etapa(et)),
            })
        st.dataframe(pd.DataFrame(resumo), use_container_width=True, hide_index=True)

        # ── HISTOGRAMA DE DIMENSIONAMENTO ──
        st.markdown("---")
        st.markdown("## Histograma de Dimensionamento de Equipe")

        hist = histograma_por_dia()
        hist_func = histograma_por_funcao()

        if hist:
            import plotly.graph_objects as go

            # Histograma empilhado por função
            funcoes_set = set()
            for d, funcs in hist_func.items():
                funcoes_set.update(funcs.keys())
            funcoes_list = sorted(funcoes_set)

            datas_sorted = sorted(hist_func.keys())
            fig = go.Figure()

            cores_funcoes = [
                "#ed1c24", "#a01e21", "#333333", "#666666", "#999999",
                "#cc3333", "#ff6666", "#993333", "#cc6666", "#ff9999",
                "#663333", "#996666", "#cc9999", "#444444", "#777777",
            ]

            for i, funcao in enumerate(funcoes_list):
                y_vals = [hist_func.get(d, {}).get(funcao, 0) for d in datas_sorted]
                cor = cores_funcoes[i % len(cores_funcoes)]
                fig.add_trace(go.Bar(
                    x=[d.strftime("%d/%m") for d in datas_sorted],
                    y=y_vals,
                    name=funcao,
                    marker_color=cor,
                ))

            fig.update_layout(
                barmode="stack",
                title="Histograma de Efetivo por Dia (Empilhado por Função)",
                xaxis_title="Data",
                yaxis_title="Quantidade de Pessoas",
                legend_title="Funções",
                height=450,
                template="plotly_white",
                font=dict(family="Arial", size=12),
            )
            st.plotly_chart(fig, use_container_width=True)

            # Métricas do histograma
            hm1, hm2, hm3, hm4 = st.columns(4)
            with hm1:
                st.metric("Dias com Equipe", f"{len(hist)}")
            with hm2:
                st.metric("Pico de Efetivo", f"{max(hist.values())} pessoas")
            with hm3:
                media_ef = sum(hist.values()) / len(hist)
                st.metric("Média de Efetivo", f"{media_ef:.1f} pessoas")
            with hm4:
                st.metric("Total DH (Histograma)", f"{sum(hist.values()):,}")

            # Tabela do histograma
            with st.expander("Tabela Detalhada do Histograma", expanded=False):
                hist_df = pd.DataFrame([
                    {"Data": d.strftime("%d/%m/%Y"), "Dia da Semana": d.strftime("%A"), "Efetivo": qtd}
                    for d, qtd in hist.items()
                ])
                st.dataframe(hist_df, use_container_width=True, hide_index=True)
        else:
            st.info("Adicione etapas com datas de início para gerar o histograma de dimensionamento.")


# ══════════════════════════════════════════════
# PÁGINA 5 — CUSTOS INDIRETOS
# ══════════════════════════════════════════════

elif pagina == "5. Custos Indiretos":
    st.markdown("# Custos Indiretos (Logística)")
    st.markdown("Todos os custos são calculados com base no cronograma de alocação e histograma de efetivo.")
    st.markdown("---")

    dh = total_diarias_homem()
    dias = total_dias_obra()

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Total de Dias", f"{dias}")
    with m2:
        st.metric("Pico de Efetivo", f"{pico_efetivo()} pessoas")
    with m3:
        st.metric("Total Diárias-Homem", f"{dh:,.0f}")

    alertas = validacoes_consistencia()
    if alertas:
        for a in alertas:
            st.markdown(f'<div class="alerta-consistencia">⚠️ {a}</div>', unsafe_allow_html=True)

    st.markdown("---")

    st.markdown("## Refeição")
    rf1, rf2 = st.columns(2)
    with rf1:
        st.session_state.refeicao_valor = st.number_input("Valor por Refeição/Dia (R$)", min_value=0.0, value=st.session_state.refeicao_valor, step=5.0)
    with rf2:
        st.metric("Total Refeição", fmt(st.session_state.refeicao_valor * dh))

    st.markdown("---")

    st.markdown("## Hospedagem")
    st.session_state.tipo_hospedagem = st.radio("Tipo de Hospedagem", ["Hotel", "Alojamento próprio", "Local (sem hospedagem)"], horizontal=True, index=["Hotel", "Alojamento próprio", "Local (sem hospedagem)"].index(st.session_state.tipo_hospedagem))

    if st.session_state.tipo_hospedagem == "Hotel":
        ho1, ho2 = st.columns(2)
        with ho1:
            st.session_state.hospedagem_valor = st.number_input("Valor Diária Hotel (R$)", min_value=0.0, value=st.session_state.hospedagem_valor, step=10.0)
        with ho2:
            st.metric("Total Hospedagem", fmt(st.session_state.hospedagem_valor * dh))
    elif st.session_state.tipo_hospedagem == "Alojamento próprio":
        st.info("Custo de alojamento contemplado no benefício/ajuda de custo dos recursos alojados.")
        st.session_state.hospedagem_valor = 0.0
    else:
        st.info("Sem custo de hospedagem — equipe local.")
        st.session_state.hospedagem_valor = 0.0

    st.markdown("---")

    st.markdown("## Mobilização e Desmobilização")
    md1, md2 = st.columns(2)
    with md1:
        st.session_state.mobilizacao_valor = st.number_input("Custo Mobilização (R$)", min_value=0.0, value=st.session_state.mobilizacao_valor, step=500.0)
    with md2:
        st.session_state.desmobilizacao_valor = st.number_input("Custo Desmobilização (R$)", min_value=0.0, value=st.session_state.desmobilizacao_valor, step=500.0)

    st.markdown("---")

    st.markdown("## Diárias de Veículos")
    vv1, vv2, vv3 = st.columns(3)
    with vv1:
        st.session_state.qtd_veiculos = st.number_input("Quantidade de Veículos", min_value=0, value=st.session_state.qtd_veiculos, step=1)
    with vv2:
        st.session_state.diaria_veiculo_valor = st.number_input("Diária por Veículo (R$)", min_value=0.0, value=st.session_state.diaria_veiculo_valor, step=50.0)
    with vv3:
        st.metric("Total Veículos", fmt(st.session_state.diaria_veiculo_valor * st.session_state.qtd_veiculos * dias))

    st.markdown("---")

    st.markdown("## Combustível e Pedágio")
    cc1, cc2, cc3 = st.columns(3)
    with cc1:
        st.session_state.distancia_km = st.number_input("Distância Base ↔ Obra (km, só ida)", min_value=0.0, value=st.session_state.distancia_km, step=10.0)
        st.session_state.qtd_viagens_mob = st.number_input("Nº de Viagens (ida+volta)", min_value=1, value=st.session_state.qtd_viagens_mob, step=1)
    with cc2:
        st.session_state.consumo_km_litro = st.number_input("Consumo (km/litro)", min_value=0.1, value=st.session_state.consumo_km_litro, step=0.5)
        st.session_state.preco_combustivel = st.number_input("Preço Combustível (R$/litro)", min_value=0.0, value=st.session_state.preco_combustivel, step=0.10)
    with cc3:
        dist = st.session_state.distancia_km
        consumo = st.session_state.consumo_km_litro
        viagens = st.session_state.qtd_viagens_mob
        litros = (dist * 2 * viagens) / consumo if consumo > 0 else 0
        custo_comb = litros * st.session_state.preco_combustivel
        st.metric("Litros Estimados", f"{litros:,.1f} L")
        st.metric("Total Combustível", fmt(custo_comb))

    st.session_state.credito_combustivel = st.checkbox("Combustível gera crédito PIS/COFINS?", value=st.session_state.credito_combustivel, help="No Lucro Real, combustível usado na atividade gera crédito")

    st.session_state.tem_pedagio = st.checkbox("Tem pedágio no percurso?", value=st.session_state.tem_pedagio)
    if st.session_state.tem_pedagio:
        pp1, pp2 = st.columns(2)
        with pp1:
            st.session_state.valor_pedagio_trecho = st.number_input("Valor Pedágio por Trecho (R$, só ida)", min_value=0.0, value=st.session_state.valor_pedagio_trecho, step=5.0)
        with pp2:
            ped_total = st.session_state.valor_pedagio_trecho * 2 * viagens
            st.metric("Total Pedágio", fmt(ped_total))

    st.markdown("---")

    st.markdown("## Composição dos Custos Indiretos")
    ci = custo_indireto_total()
    ci_df = pd.DataFrame([
        {"Item": "Refeição", "Valor": fmt(ci["refeicao"])},
        {"Item": "Hospedagem", "Valor": fmt(ci["hospedagem"])},
        {"Item": "Mobilização + Desmobilização", "Valor": fmt(ci["mobilizacao"])},
        {"Item": "Veículos", "Valor": fmt(ci["veiculos"])},
        {"Item": "Combustível", "Valor": fmt(ci["combustivel"])},
        {"Item": "Pedágio", "Valor": fmt(ci["pedagio"])},
        {"Item": "TOTAL", "Valor": fmt(ci["total"])},
    ])
    st.table(ci_df)


# ══════════════════════════════════════════════
# PÁGINA 6 — FECHAMENTO
# ══════════════════════════════════════════════

elif pagina == "6. Fechamento":
    st.markdown(f"# Fechamento — {st.session_state.nome_obra or 'Obra sem nome'}")
    st.markdown("DRE Gerencial, BDI reverso, Simulador de Negociação e Relatórios (Lucro Real).")
    st.markdown("---")

    recalcular_custos_hora()

    cd_mo = custo_direto_mo()
    cd_eq = custo_direto_equip()
    cd_mat = custo_direto_materiais_bruto()
    cd = custo_direto_total()
    ci = custo_indireto_total()
    creditos = total_creditos_tributarios()
    risco_val = custo_total_bruto() * (st.session_state.markup_risco / 100.0)
    ct_bruto = custo_total_bruto()
    ct_liq = custo_total_liquido()
    pv = preco_venda()
    imp_fat_pct = impostos_faturamento_pct()
    imp_valor = pv * (imp_fat_pct / 100.0)
    lucro_bruto_p = lucro_bruto_pct_efetivo()
    lucro_valor = pv * (lucro_bruto_p / 100.0)
    rl = pv - imp_valor

    meses_obra = max(total_dias_obra() / 30, 1)
    irpj = calcular_irpj_csll(lucro_valor / meses_obra, meses_obra) if lucro_valor > 0 else {"irpj_base": 0, "irpj_adicional": 0, "irpj_total": 0, "csll": 0, "total": 0, "lucro_total": 0}
    lucro_liquido = lucro_valor - irpj["total"]

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Custo Direto", fmt(cd))
    with k2:
        st.metric("Custo Indireto", fmt(ci["total"]))
    with k3:
        st.metric("Custo Total Líquido", fmt(ct_liq))
    with k4:
        st.metric("Preço de Venda", fmt(pv))

    k5, k6, k7, k8 = st.columns(4)
    with k5:
        st.metric("Créditos PIS/COFINS", fmt(creditos))
    with k6:
        st.metric("Lucro Bruto", fmt(lucro_valor))
    with k7:
        st.metric("IRPJ + CSLL", fmt(irpj["total"]))
    with k8:
        st.metric("Lucro Líquido", fmt(lucro_liquido))

    st.markdown("---")

    # ── DRE ──
    st.markdown("## DRE Gerencial")
    if pv > 0:
        dre = [
            {"Descrição": "Receita Bruta (Preço de Venda)", "Valor (R$)": fmt(pv), "% s/ PV": "100,00%"},
            {"Descrição": f"(-) PIS ({st.session_state.pis_pct:.2f}%)", "Valor (R$)": f"-{fmt(pv * st.session_state.pis_pct / 100)}", "% s/ PV": f"-{st.session_state.pis_pct:.2f}%"},
            {"Descrição": f"(-) COFINS ({st.session_state.cofins_pct:.2f}%)", "Valor (R$)": f"-{fmt(pv * st.session_state.cofins_pct / 100)}", "% s/ PV": f"-{st.session_state.cofins_pct:.2f}%"},
            {"Descrição": f"(-) ISS ({st.session_state.iss_pct:.2f}%)", "Valor (R$)": f"-{fmt(pv * st.session_state.iss_pct / 100)}", "% s/ PV": f"-{st.session_state.iss_pct:.2f}%"},
            {"Descrição": "(=) Receita Líquida", "Valor (R$)": fmt(rl), "% s/ PV": f"{rl/pv*100:.2f}%"},
            {"Descrição": "(-) Custo Direto — Mão de Obra", "Valor (R$)": f"-{fmt(cd_mo)}", "% s/ PV": f"-{cd_mo/pv*100:.2f}%"},
            {"Descrição": "(-) Custo Direto — Equipamentos", "Valor (R$)": f"-{fmt(cd_eq)}", "% s/ PV": f"-{cd_eq/pv*100:.2f}%"},
            {"Descrição": "(-) Custo Direto — Materiais", "Valor (R$)": f"-{fmt(cd_mat)}", "% s/ PV": f"-{cd_mat/pv*100:.2f}%"},
            {"Descrição": "(+) Créditos PIS/COFINS (insumos)", "Valor (R$)": f"+{fmt(creditos)}", "% s/ PV": f"+{creditos/pv*100:.2f}%"},
            {"Descrição": "(-) Custos Indiretos (Logística)", "Valor (R$)": f"-{fmt(ci['total'])}", "% s/ PV": f"-{ci['total']/pv*100:.2f}%"},
            {"Descrição": f"(-) Markup de Risco ({st.session_state.markup_risco:.1f}%)", "Valor (R$)": f"-{fmt(risco_val)}", "% s/ PV": f"-{risco_val/pv*100:.2f}%"},
            {"Descrição": "(=) Lucro antes IR/CSLL", "Valor (R$)": fmt(lucro_valor), "% s/ PV": f"{lucro_bruto_p:.2f}%"},
            {"Descrição": "(-) IRPJ (15% + 10% adic.)", "Valor (R$)": f"-{fmt(irpj['irpj_total'])}", "% s/ PV": f"-{irpj['irpj_total']/pv*100:.2f}%"},
            {"Descrição": "(-) CSLL (9%)", "Valor (R$)": f"-{fmt(irpj['csll'])}", "% s/ PV": f"-{irpj['csll']/pv*100:.2f}%"},
            {"Descrição": "(=) LUCRO LÍQUIDO", "Valor (R$)": fmt(lucro_liquido), "% s/ PV": f"{lucro_liquido/pv*100:.2f}%"},
        ]
        st.table(pd.DataFrame(dre))

        if st.session_state.modo_lucro == "Lucro Líquido Alvo (Gross-up)":
            st.info(f"Modo Gross-up ativo: Lucro líquido alvo de **{st.session_state.lucro_pct:.2f}%** → Lucro bruto calculado de **{lucro_bruto_p:.2f}%** para compensar IRPJ/CSLL.")
    else:
        st.warning("Sem dados suficientes para gerar o DRE.")

    st.markdown("---")

    # ── IRPJ detalhado ──
    st.markdown("## IRPJ e CSLL — Cálculo Progressivo")
    if pv > 0 and lucro_valor > 0:
        ir1, ir2 = st.columns(2)
        with ir1:
            st.markdown(f"""
| Item | Valor |
|------|-------|
| Lucro da Obra | {fmt(lucro_valor)} |
| Duração estimada | {meses_obra:.1f} meses |
| Lucro mensal médio | {fmt(lucro_valor / meses_obra)} |
| Limite IRPJ 15% | R$ 20.000,00/mês |
| Excedente mensal | {fmt(max(0, lucro_valor / meses_obra - 20000))} |
            """)
        with ir2:
            st.markdown(f"""
| Tributo | Valor |
|---------|-------|
| IRPJ Base (15%) | {fmt(irpj['irpj_base'])} |
| IRPJ Adicional (10%) | {fmt(irpj['irpj_adicional'])} |
| CSLL (9%) | {fmt(irpj['csll'])} |
| **Total IRPJ + CSLL** | **{fmt(irpj['total'])}** |
            """)

    st.markdown("---")

    # ── BDI ──
    st.markdown("## Formação do Preço de Venda (BDI Reverso)")
    bdi_div = 1 - imp_fat_pct / 100 - lucro_bruto_p / 100

    bf1, bf2 = st.columns(2)
    with bf1:
        st.markdown(f"""
| Componente | Valor |
|------------|-------|
| Custo Direto (MO) | {fmt(cd_mo)} |
| Custo Direto (Equip.) | {fmt(cd_eq)} |
| Custo Direto (Materiais) | {fmt(cd_mat)} |
| Custos Indiretos | {fmt(ci['total'])} |
| Custo Bruto Total | {fmt(ct_bruto)} |
| (-) Créditos PIS/COFINS | -{fmt(creditos)} |
| (+) Markup de Risco | +{fmt(risco_val)} |
| **Custo Total Líquido** | **{fmt(ct_liq)}** |
| Impostos s/ Faturamento | {fmt_pct(imp_fat_pct)} |
| Lucro Bruto (BDI) | {fmt_pct(lucro_bruto_p)} |
| Divisor BDI | {bdi_div:.4f} |
| **Preço de Venda** | **{fmt(pv)}** |
        """)
    with bf2:
        st.markdown("**Fórmula:**")
        st.latex(r"\text{PV} = \frac{\text{Custo Total Líquido}}{1 - \%\text{Impostos} - \%\text{Lucro}}")
        if ct_liq > 0:
            bdi_markup = ((pv - ct_liq) / ct_liq) * 100
            st.markdown(f"**BDI (Markup):** {bdi_markup:.2f}%")
        st.markdown(f"**PV** = {fmt(ct_liq)} / {bdi_div:.4f} = **{fmt(pv)}**")

    st.markdown("---")

    # ── SIMULADOR DE NEGOCIAÇÃO ──
    st.markdown("## Simulador de Negociação Comercial")
    st.markdown("Simule descontos e veja o impacto direto na margem de lucro.")

    st.markdown('<div class="simulador-desconto">', unsafe_allow_html=True)
    desc_pct = st.slider("Desconto Comercial (%)", min_value=0.0, max_value=30.0, value=st.session_state.desconto_comercial_pct, step=0.5, key="slider_desconto")
    st.session_state.desconto_comercial_pct = desc_pct

    if pv > 0:
        sim = simular_desconto(desc_pct)
        sc1, sc2, sc3, sc4 = st.columns(4)
        with sc1:
            st.metric("PV Original", fmt(sim["pv_original"]))
        with sc2:
            st.metric("PV com Desconto", fmt(sim["pv_desconto"]), delta=f"-{fmt_pct(desc_pct)}")
        with sc3:
            st.metric("Lucro Líquido", fmt(sim["lucro_liquido"]),
                       delta=f"{sim['margem_liq_pct']:.2f}%",
                       delta_color="normal" if sim["lucro_liquido"] > 0 else "inverse")
        with sc4:
            if sim["breakeven"]:
                st.error("ABAIXO DO PONTO DE EQUILÍBRIO")
            else:
                st.success(f"Margem Líquida: {sim['margem_liq_pct']:.2f}%")

        # Tabela de cenários
        st.markdown("### Cenários de Desconto")
        cenarios = []
        for d in [0, 2, 5, 8, 10, 12, 15, 20]:
            s = simular_desconto(d)
            cenarios.append({
                "Desconto": f"{d}%",
                "PV": fmt(s["pv_desconto"]),
                "Lucro Bruto": fmt(s["lucro_bruto"]),
                "IRPJ+CSLL": fmt(s["irpj_csll"]),
                "Lucro Líquido": fmt(s["lucro_liquido"]),
                "Margem Líq.": f"{s['margem_liq_pct']:.2f}%",
                "Status": "INVIÁVEL" if s["breakeven"] else "OK",
            })
        st.dataframe(pd.DataFrame(cenarios), use_container_width=True, hide_index=True)

        # Desconto máximo
        desc_max = 0.0
        for d_test in [x * 0.5 for x in range(0, 61)]:
            s_test = simular_desconto(d_test)
            if s_test["breakeven"]:
                desc_max = max(0, d_test - 0.5)
                break
            desc_max = d_test
        st.info(f"Desconto máximo antes do ponto de equilíbrio: **{desc_max:.1f}%** (PV mínimo: **{fmt(preco_venda_com_desconto(desc_max))}**)")

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")

    # ── Detalhamento por etapa ──
    st.markdown("## Detalhamento por Etapa")
    if st.session_state.etapas:
        det = []
        for i, et in enumerate(st.session_state.etapas):
            pessoas = sum(a["qtd"] for a in et.get("alocacoes", []))
            ce = custo_etapa(et)
            det.append({
                "Etapa": f"{i+1}. {et['nome']}",
                "Data": et.get("data_inicio", "N/I"),
                "Dias": et["dias"], "Efetivo": pessoas,
                "DH": et["dias"] * pessoas,
                "Custo MO": fmt(custo_mo_etapa(et)),
                "Custo Equip": fmt(custo_equip_etapa(et)),
                "Custo Mat": fmt(custo_materiais_etapa(et)),
                "Total": fmt(ce),
                "% CD": f"{ce/cd*100:.1f}%" if cd > 0 else "0%",
            })
        st.dataframe(pd.DataFrame(det), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Relatórios ──
    st.markdown("## Relatório Final")
    st.caption("Gere o relatório completo para impressão, PDF, Excel ou compartilhamento com IA.")

    r1, r2, r3 = st.columns(3)

    with r1:
        if st.button("Gerar Relatório PDF", use_container_width=True):
            with st.spinner("Gerando PDF..."):
                md_content = gerar_relatorio_markdown()
                pdf_bytes = gerar_pdf_bytes(md_content)
                st.session_state["pdf_bytes"] = pdf_bytes
                st.session_state["md_content"] = md_content
                st.success("PDF gerado!")

    with r2:
        if st.button("Gerar Excel", use_container_width=True):
            with st.spinner("Gerando Excel..."):
                excel_bytes = gerar_excel_bytes()
                st.session_state["excel_bytes"] = excel_bytes
                st.success("Excel gerado!")

    with r3:
        if st.button("Gerar Markdown (IA)", use_container_width=True):
            md_content = gerar_relatorio_markdown()
            st.session_state["md_content"] = md_content
            st.success("Markdown gerado!")

    # Downloads
    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        if "pdf_bytes" in st.session_state:
            nome_arq = f"Orcamento_{st.session_state.nome_obra or 'Obra'}_{datetime.now().strftime('%Y%m%d')}.pdf".replace(" ", "_")
            st.download_button("Baixar PDF", data=st.session_state["pdf_bytes"], file_name=nome_arq, mime="application/pdf", use_container_width=True)
    with dl2:
        if "excel_bytes" in st.session_state:
            nome_xls = f"Orcamento_{st.session_state.nome_obra or 'Obra'}_{datetime.now().strftime('%Y%m%d')}.xlsx".replace(" ", "_")
            st.download_button("Baixar Excel", data=st.session_state["excel_bytes"], file_name=nome_xls, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with dl3:
        if "md_content" in st.session_state:
            nome_md = f"Orcamento_{st.session_state.nome_obra or 'Obra'}_{datetime.now().strftime('%Y%m%d')}.md".replace(" ", "_")
            st.download_button("Baixar Markdown", data=st.session_state["md_content"], file_name=nome_md, mime="text/markdown", use_container_width=True)

    if "md_content" in st.session_state:
        with st.expander("Relatório Markdown (copie e cole no ChatGPT/Gemini)", expanded=False):
            st.code(st.session_state["md_content"], language="markdown")

    st.markdown("---")
    if pv <= 0:
        st.error("Não foi possível calcular o Preço de Venda. Verifique se Impostos + Lucro < 100%.")
    elif ct_liq == 0:
        st.warning("O Custo Total é zero. Cadastre etapas no cronograma.")
    else:
        st.success(f"Orçamento calculado. Preço de Venda: **{fmt(pv)}** | Lucro Líquido: **{fmt(lucro_liquido)}**")


# ══════════════════════════════════════════════
# PÁGINA 7 — BANCO DE DADOS
# ══════════════════════════════════════════════

elif pagina == "7. Banco de Dados":
    st.markdown("# Banco de Dados de Orçamentos")
    st.markdown("Salve, carregue e versione seus orçamentos. Os dados são persistidos em SQLite.")
    st.markdown("---")

    from database import salvar_orcamento, listar_orcamentos, carregar_orcamento, listar_versoes, carregar_versao, excluir_orcamento

    # Salvar
    st.markdown("## Salvar Orçamento Atual")
    sv1, sv2 = st.columns(2)
    with sv1:
        if st.button("Salvar como Novo Orçamento", use_container_width=True):
            oid = salvar_orcamento(state_to_dict())
            st.session_state.orcamento_id = oid
            st.success(f"Orçamento salvo com ID #{oid}")
    with sv2:
        if st.session_state.orcamento_id:
            if st.button(f"Atualizar Orçamento #{st.session_state.orcamento_id}", use_container_width=True):
                salvar_orcamento(state_to_dict(), st.session_state.orcamento_id)
                st.success(f"Orçamento #{st.session_state.orcamento_id} atualizado (nova versão)!")
        else:
            st.info("Salve primeiro para poder atualizar.")

    st.markdown("---")

    # Listar
    st.markdown("## Orçamentos Salvos")
    orcamentos = listar_orcamentos()
    if orcamentos:
        df_orc = pd.DataFrame(orcamentos)
        df_orc = df_orc.rename(columns={
            "id": "ID", "nome": "Obra", "cliente": "Cliente",
            "local_obra": "Local", "versao": "Versão",
            "status": "Status", "criado_em": "Criado em",
            "atualizado_em": "Atualizado em"
        })
        st.dataframe(df_orc, use_container_width=True, hide_index=True)

        st.markdown("### Carregar Orçamento")
        orc_ids = {o["id"]: f"#{o['id']} — {o['nome']} (v{o['versao']})" for o in orcamentos}
        sel_orc = st.selectbox("Selecione o orçamento", options=list(orc_ids.keys()), format_func=lambda x: orc_ids[x])

        lc1, lc2, lc3 = st.columns(3)
        with lc1:
            if st.button("Carregar", use_container_width=True):
                dados = carregar_orcamento(sel_orc)
                if dados:
                    dict_to_state(dados)
                    st.session_state.orcamento_id = sel_orc
                    st.success(f"Orçamento #{sel_orc} carregado!")
                    st.rerun()
                else:
                    st.error("Erro ao carregar orçamento.")
        with lc2:
            versoes = listar_versoes(sel_orc)
            if versoes:
                ver_ids = {v["id"]: f"Versão {v['versao']} — {v['criado_em'][:16]}" for v in versoes}
                sel_ver = st.selectbox("Versão anterior", options=list(ver_ids.keys()), format_func=lambda x: ver_ids[x])
                if st.button("Restaurar Versão", use_container_width=True):
                    dados = carregar_versao(sel_ver)
                    if dados:
                        dict_to_state(dados)
                        st.session_state.orcamento_id = sel_orc
                        st.success("Versão restaurada!")
                        st.rerun()
            else:
                st.info("Sem versões anteriores.")
        with lc3:
            if st.button("Excluir Orçamento", type="secondary", use_container_width=True):
                excluir_orcamento(sel_orc)
                if st.session_state.orcamento_id == sel_orc:
                    st.session_state.orcamento_id = None
                st.success(f"Orçamento #{sel_orc} excluído.")
                st.rerun()
    else:
        st.info("Nenhum orçamento salvo ainda. Use o botão acima para salvar o orçamento atual.")
